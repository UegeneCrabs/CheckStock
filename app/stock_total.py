from __future__ import annotations

import io
from collections import defaultdict
from datetime import datetime

from app.domain import MOSCOW_TIMEZONE
from app.repositories import stock_total as repository
from app.stores import STORES

MARKETPLACES = (
    ("WB", "wb", "ВБ"),
    ("OZON", "ozon", "ОЗОН"),
    ("YANDEX MARKET", "yandex", "ЯМ"),
)
STOCK_GROUPS = (
    ("ff", "ДОСТУПНО ФФ ДЛЯ РАСПРЕДЕЛЕНИЯ"),
    ("transit", "В ПУТИ МЕЖДУ ФФ"),
    ("fbs", "ТЕКУЩИЙ СТОК В ПРОДАЖЕ FBS"),
    ("rfbs", "ТЕКУЩИЙ СТОК В ПРОДАЖЕ RFBS"),
    ("fbo", "ТЕКУЩИЙ СТОК В ПРОДАЖЕ FBO"),
)
QUANTITY_KEYS = tuple(
    f"{group}_{marketplace_key}"
    for group, _group_label in STOCK_GROUPS
    for _marketplace, marketplace_key, _marketplace_label in MARKETPLACES
)
MARKETPLACE_TOTAL_KEYS = tuple(
    f"total_{marketplace_key}"
    for _marketplace, marketplace_key, _marketplace_label in MARKETPLACES
)
TOTAL_KEYS = ("grand_total", *MARKETPLACE_TOTAL_KEYS)
VALUE_KEYS = (*TOTAL_KEYS, *QUANTITY_KEYS)


def _normalized(value: object) -> str:
    return str(value or "").replace("\xa0", " ").strip().casefold()


def _scheme_group(scheme: object) -> str | None:
    normalized = _normalized(scheme)
    if normalized == "fbs" or normalized.startswith("fbs_"):
        return "fbs"
    if normalized in {"rfbs", "fbo"}:
        return normalized
    return None


def _empty_row(store_slug: str, identity: tuple[str, ...]) -> dict:
    row = {
        "store_slug": store_slug,
        "store_name": STORES.get(store_slug, {}).get("name", store_slug.upper()),
        "identity": identity,
        "article": "",
        "barcode": "",
        "name": "",
        "purchase_price": None,
        "grand_total": 0,
    }
    row.update({key: 0 for key in MARKETPLACE_TOTAL_KEYS})
    row.update({key: 0 for key in QUANTITY_KEYS})
    return row


def build_rows(
    store_slugs: tuple[str, ...],
    allowed_pairs: tuple[tuple[str, str], ...] | None = None,
) -> list[dict]:
    catalog, marketplace_stock, fulfillment_stock, transit_stock = repository.get_source_rows(
        store_slugs
    )
    if allowed_pairs is not None:
        allowed = set(allowed_pairs)
        catalog = [
            item
            for item in catalog
            if (str(item["store_slug"]), str(item["marketplace"])) in allowed
        ]
        marketplace_stock = [
            item
            for item in marketplace_stock
            if (str(item["store_slug"]), str(item["marketplace"])) in allowed
        ]
        fulfillment_stock = [
            item
            for item in fulfillment_stock
            if (str(item["store_slug"]), str(item["marketplace"])) in allowed
        ]
        transit_stock = [
            item
            for item in transit_stock
            if (str(item["store_slug"]), str(item["marketplace"])) in allowed
        ]
    marketplace_keys = {marketplace: key for marketplace, key, _label in MARKETPLACES}

    barcode_identity_by_article: dict[tuple[str, str], set[tuple[str, ...]]] = defaultdict(set)
    for item in catalog:
        barcode = _normalized(item.get("barcode"))
        article = _normalized(item.get("article"))
        if barcode and article:
            barcode_identity_by_article[(str(item["store_slug"]), article)].add(("barcode", barcode))

    identity_by_offer: dict[tuple[str, str, str], tuple[str, ...]] = {}
    buckets: dict[tuple[str, tuple[str, ...]], dict] = {}

    for item in catalog:
        store_slug = str(item["store_slug"])
        marketplace = str(item["marketplace"])
        article = str(item.get("article") or "").strip()
        article_key = _normalized(article)
        barcode = str(item.get("barcode") or "").strip()
        barcode_key = _normalized(barcode)
        if barcode_key:
            identity = ("barcode", barcode_key)
        else:
            candidates = barcode_identity_by_article.get((store_slug, article_key), set())
            identity = (
                next(iter(candidates))
                if len(candidates) == 1
                else (
                    "offer",
                    marketplace,
                    article_key,
                )
            )
        identity_by_offer[(store_slug, marketplace, article)] = identity
        bucket_key = (store_slug, identity)
        row = buckets.setdefault(bucket_key, _empty_row(store_slug, identity))
        if not row["article"]:
            row["article"] = article
        if not row["barcode"] and barcode:
            row["barcode"] = barcode
        name = str(item.get("name") or "").strip()
        if not row["name"] and name:
            row["name"] = name
        purchase_price = item.get("purchase_price")
        if row["purchase_price"] is None and purchase_price is not None:
            row["purchase_price"] = float(purchase_price)

    identities_by_article: dict[tuple[str, str], set[tuple[str, ...]]] = defaultdict(set)
    for (store_slug, _marketplace, article), identity in identity_by_offer.items():
        identities_by_article[(store_slug, _normalized(article))].add(identity)

    def identity_for(store_slug: str, marketplace: str, article: str) -> tuple[str, ...]:
        exact = identity_by_offer.get((store_slug, marketplace, article))
        if exact is not None:
            return exact
        candidates = identities_by_article.get((store_slug, _normalized(article)), set())
        if len(candidates) == 1:
            return next(iter(candidates))
        return ("orphan", marketplace, _normalized(article))

    def bucket_for(store_slug: str, marketplace: str, article: str) -> dict:
        identity = identity_for(store_slug, marketplace, article)
        row = buckets.setdefault((store_slug, identity), _empty_row(store_slug, identity))
        if not row["article"]:
            row["article"] = article
        if not row["name"]:
            row["name"] = article
        return row

    for item in fulfillment_stock:
        marketplace = str(item["marketplace"])
        marketplace_key = marketplace_keys.get(marketplace)
        if marketplace_key is None:
            continue
        store_slug = str(item["store_slug"])
        article = str(item["article"])
        row = bucket_for(store_slug, marketplace, article)
        row[f"ff_{marketplace_key}"] += int(item.get("quantity") or 0)

    for item in transit_stock:
        marketplace = str(item["marketplace"])
        marketplace_key = marketplace_keys.get(marketplace)
        if marketplace_key is None:
            continue
        store_slug = str(item["store_slug"])
        article = str(item["article"])
        row = bucket_for(store_slug, marketplace, article)
        row[f"transit_{marketplace_key}"] += int(item.get("quantity") or 0)

    for item in marketplace_stock:
        marketplace = str(item["marketplace"])
        marketplace_key = marketplace_keys.get(marketplace)
        group = _scheme_group(item.get("scheme"))
        if marketplace_key is None or group is None:
            continue
        store_slug = str(item["store_slug"])
        article = str(item["article"])
        row = bucket_for(store_slug, marketplace, article)
        row[f"{group}_{marketplace_key}"] += int(item.get("quantity") or 0)

    rows = list(buckets.values())
    for row in rows:
        for _marketplace, marketplace_key, _marketplace_label in MARKETPLACES:
            row[f"total_{marketplace_key}"] = sum(
                int(row[f"{group}_{marketplace_key}"] or 0)
                for group, _group_label in STOCK_GROUPS
            )
        row["grand_total"] = sum(int(row[key] or 0) for key in MARKETPLACE_TOTAL_KEYS)
        row.pop("identity", None)
    rows.sort(
        key=lambda row: (
            -int(row["grand_total"]),
            str(row["store_name"]).casefold(),
            str(row["name"]).casefold(),
            str(row["article"]).casefold(),
        )
    )
    return rows


def build_xlsx(rows: list[dict]) -> tuple[bytes, str]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "для выгрузки в .xlsx нужен пакет openpyxl — установи его в .venv (pip install openpyxl)"
        ) from error

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Остатки Тотал"
    sheet.sheet_view.showGridLines = False

    fixed_headers = ("МАГАЗИН", "АРТИКУЛ", "ШТРИХКОД", "НАЗВАНИЕ")
    for column, title in enumerate(fixed_headers, start=1):
        sheet.cell(row=1, column=column, value=title)
        sheet.merge_cells(start_row=1, start_column=column, end_row=2, end_column=column)

    group_fills = {
        "total": "E4E7F7",
        "ff": "FFF2CC",
        "transit": "FCE4D6",
        "fbs": "DDEBF7",
        "rfbs": "E4DFEC",
        "fbo": "E2F0D9",
    }
    column = len(fixed_headers) + 1
    total_start = column
    total_end = column + len(TOTAL_KEYS) - 1
    sheet.cell(row=1, column=total_start, value="ТОТАЛ")
    sheet.merge_cells(start_row=1, start_column=total_start, end_row=1, end_column=total_end)
    for offset, label in enumerate(("ГРАНД ТОТАЛ", "ВБ", "ОЗОН", "ЯМ")):
        sheet.cell(row=2, column=total_start + offset, value=label)
    for row_number in (1, 2):
        for group_column in range(total_start, total_end + 1):
            sheet.cell(row=row_number, column=group_column).fill = PatternFill(
                "solid", fgColor=group_fills["total"]
            )
    column = total_end + 1
    for group, label in STOCK_GROUPS:
        start = column
        end = column + len(MARKETPLACES) - 1
        sheet.cell(row=1, column=start, value=label)
        sheet.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        for offset, (_marketplace, _key, marketplace_label) in enumerate(MARKETPLACES):
            sheet.cell(row=2, column=start + offset, value=marketplace_label)
        for row_number in (1, 2):
            for group_column in range(start, end + 1):
                sheet.cell(row=row_number, column=group_column).fill = PatternFill(
                    "solid", fgColor=group_fills[group]
                )
        column = end + 1

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B7C3D0")
    medium = Side(style="medium", color="7F8C9A")
    for row_number in (1, 2):
        for cell in sheet[row_number]:
            if cell.column <= len(fixed_headers):
                cell.fill = header_fill
                cell.font = header_font
            else:
                cell.font = Font(bold=True, color="172033")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=medium)

    total_values = ["ИТОГО", "", "", f"позиций: {len(rows)}"]
    total_values.extend(sum(int(row.get(key) or 0) for row in rows) for key in VALUE_KEYS)
    sheet.append(total_values)
    priced_positions = sum(1 for row in rows if row.get("purchase_price") is not None)
    cost_values = ["ИТОГО В ЗЦ", "", "", f"ЗЦ: {priced_positions} из {len(rows)} поз."]
    cost_values.extend(
        round(
            sum(
                float(row.get("purchase_price") or 0) * int(row.get(key) or 0)
                for row in rows
                if row.get("purchase_price") is not None
            ),
            2,
        )
        for key in VALUE_KEYS
    )
    sheet.append(cost_values)
    for total_row_number, fill_color in ((3, "FFF4E5"), (4, "EAF4EE")):
        for cell in sheet[total_row_number]:
            cell.font = Font(bold=True, color="172033")
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, bottom=medium)

    for row in rows:
        sheet.append(
            [
                row["store_name"],
                str(row["article"]),
                str(row["barcode"]),
                row["name"],
                *(int(row.get(key) or 0) for key in VALUE_KEYS),
            ]
        )

    for row_number in range(5, sheet.max_row + 1):
        for column_number in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_number, column=column_number)
            cell.border = Border(bottom=Side(style="hair", color="D9E1E8"))
            cell.alignment = Alignment(
                horizontal="left" if column_number <= 4 else "right",
                vertical="center",
                wrap_text=column_number == 4,
            )
        sheet.cell(row=row_number, column=2).number_format = "@"
        sheet.cell(row=row_number, column=3).number_format = "@"
        for column_number in range(5, sheet.max_column + 1):
            sheet.cell(row=row_number, column=column_number).number_format = "#,##0"
    for column_number in range(5, sheet.max_column + 1):
        sheet.cell(row=3, column=column_number).number_format = "#,##0"
        sheet.cell(row=4, column=column_number).number_format = '#,##0.00 "₽"'

    widths = [18, 18, 20, 54, 16, 12, 12, 12] + [12] * len(QUANTITY_KEYS)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 38
    sheet.row_dimensions[2].height = 26
    sheet.row_dimensions[3].height = 24
    sheet.row_dimensions[4].height = 24
    sheet.freeze_panes = "I5"

    buffer = io.BytesIO()
    workbook.save(buffer)
    date_label = datetime.now(MOSCOW_TIMEZONE).strftime("%Y-%m-%d")
    return buffer.getvalue(), f"ostatki_total_{date_label}.xlsx"
