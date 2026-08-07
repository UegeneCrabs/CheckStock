"""
Сборка .xlsx по сохранённой операции со стоком.

Сами файлы, которыми пользовались при загрузке, не хранятся — вместо них
в БД лежат применённые строки (stock_operation_items). Файл собирается в
момент скачивания. Так одинаково работает и для загрузки файлом, и по
ссылке на Google Таблицу, и для ручного ввода, и база не распухает от
вложений, которые почти никогда не понадобятся.
"""

import io

from app import db
from app.formatting import format_dt


def _safe_filename(value: str) -> str:
    """Имя файла без символов, которые ломают заголовок Content-Disposition."""
    cleaned = "".join(c for c in value if c.isalnum() or c in " _-.()")
    return cleaned.strip() or "operation"


def build_operation_xlsx(operation_id: int) -> tuple[bytes, str]:
    """Возвращает (содержимое файла, имя файла) по id операции."""
    operation = db.get_operation(operation_id)
    if operation is None:
        raise LookupError("операция не найдена")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        raise RuntimeError(
            "для выгрузки в .xlsx нужен пакет openpyxl — установи его в .venv "
            "(pip install openpyxl)"
        ) from e

    items = db.get_operation_items(operation_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Позиции"
    ws.append(["BARCODE", "ARTICLE", "НАЗВАНИЕ", "КОЛИЧЕСТВО"])
    for item in items:
        ws.append([item["barcode"] or "", item["article"], item["name"] or "", item["quantity"]])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 14

    # Второй лист — обстоятельства операции: кто, когда, откуда и куда.
    # Без него по одному списку товаров не понять, что это была за выгрузка.
    info = wb.create_sheet("Операция")
    source_from = _cell_label(operation["from_fulfillment"], operation["from_marketplace"])
    source_to = _cell_label(operation["to_fulfillment"], operation["to_marketplace"])

    rows = [
        ("Тип", db.OPERATION_LABELS.get(operation["kind"], operation["kind"])),
        ("Магазин", operation["store_slug"].upper()),
        ("Источник данных", db.SOURCE_LABELS.get(operation["source_type"], operation["source_type"])),
        ("Название файла/таблицы", operation["source_name"] or "—"),
        ("Ссылка", operation["sheet_url"] or "—"),
    ]

    # У отгрузки получателя нет, у поставки — источника. Пустые строки
    # «Куда: —» только сбивают с толку, поэтому показываем то, что есть.
    if source_from != "—":
        rows.append(("Откуда" if operation["kind"] != "shipment" else "Ушло со склада", source_from))
    if source_to != "—":
        rows.append(("Куда", source_to))

    note = _operation_field(operation, "note")
    if note:
        rows.append(("Примечание", note))

    rows += [
        ("Сотрудник", operation["user_name"]),
        ("Когда", format_dt(operation["created_at"])),
        ("Позиций", len(items)),
        ("Всего единиц", sum(i["quantity"] for i in items)),
    ]
    for row_idx, (key, value) in enumerate(rows, start=1):
        info.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        info.cell(row=row_idx, column=2, value=value)
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 60

    buffer = io.BytesIO()
    wb.save(buffer)

    date_part = (operation["created_at"] or "")[:10]
    name = _safe_filename(
        f"{operation['kind']}_{operation['store_slug']}_{date_part}_{operation_id}"
    )
    return buffer.getvalue(), f"{name}.xlsx"


def _operation_field(operation: dict, name: str) -> str:
    """Поле операции, которого может не быть у старых записей: колонка note
    появилась позже, и в уже сохранённых операциях её нет."""
    try:
        return str(operation[name] or "")
    except (KeyError, IndexError):
        return ""


def _cell_label(fulfillment: str | None, marketplace: str | None) -> str:
    if not fulfillment and not marketplace:
        return "—"
    return " / ".join(part for part in (fulfillment, marketplace) if part)


def build_history_xlsx(store_slug: str, store_name: str,
                       operations: list[dict]) -> tuple[bytes, str]:
    """Вся история движений стока магазина одним файлом.

    Два листа. «Операции» — по строке на операцию, чтобы окинуть взглядом
    и отфильтровать в Excel. «Позиции» — по строке на товар, с продублированными
    полями операции: так по выгрузке можно построить сводную таблицу, а без
    дублирования пришлось бы связывать листы руками.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "для выгрузки в .xlsx нужен пакет openpyxl — установи его в .venv "
            "(pip install openpyxl)"
        ) from e

    wb = openpyxl.Workbook()

    def _head(ws, titles: list[str], widths: list[int]) -> None:
        ws.append(titles)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"

    summary = wb.active
    summary.title = "Операции"
    _head(
        summary,
        ["Когда", "Тип", "Откуда", "Куда", "Позиций", "Единиц",
         "Источник", "Название файла/таблицы", "Примечание", "Сотрудник"],
        [20, 20, 30, 30, 10, 10, 16, 32, 34, 24],
    )

    lines = wb.create_sheet("Позиции")
    _head(
        lines,
        ["Когда", "Тип", "Откуда", "Куда", "BARCODE", "ARTICLE", "НАЗВАНИЕ",
         "КОЛИЧЕСТВО", "Примечание", "Сотрудник"],
        [20, 20, 30, 30, 20, 16, 46, 13, 34, 24],
    )

    for op in operations:
        when = format_dt(op["created_at"])
        kind = db.OPERATION_LABELS.get(op["kind"], op["kind"])
        src = _cell_label(op["from_fulfillment"], op["from_marketplace"])
        dst = _cell_label(op["to_fulfillment"], op["to_marketplace"])
        note = _operation_field(op, "note")

        summary.append([
            when, kind, src, dst,
            op.get("positions", 0), op.get("units", 0),
            db.SOURCE_LABELS.get(op["source_type"], op["source_type"]),
            op["source_name"] or op["sheet_url"] or "",
            note, op["user_name"],
        ])

        for item in op.get("items", []):
            lines.append([
                when, kind, src, dst,
                item.get("barcode") or "", item.get("article", ""),
                item.get("name") or "", item.get("quantity", 0),
                note, op["user_name"],
            ])

    buffer = io.BytesIO()
    wb.save(buffer)

    name = _safe_filename(f"istoriya_stoka_{store_slug}")
    return buffer.getvalue(), f"{name}.xlsx"


def build_warehouses_xlsx(store_slug: str, store_name: str, marketplace: str,
                          tables: list[tuple[str, list[dict]]]) -> tuple[bytes, str]:
    """Детализация по складам одним файлом: лист на каждую вкладку.

    tables — [(название вкладки, строки как в get_mp_warehouse_details)].

    Раскладываем «товар × склад» в плоские строки, а не в широкую таблицу с
    колонками-складами: в Excel по плоскому списку строится любая сводная,
    а широкую таблицу сначала пришлось бы разворачивать обратно.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "для выгрузки в .xlsx нужен пакет openpyxl — установи его в .venv "
            "(pip install openpyxl)"
        ) from e

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for title, rows in tables:
        # имя листа в Excel не длиннее 31 символа и без служебных знаков
        sheet_name = _safe_filename(title)[:31] or "Лист"
        ws = wb.create_sheet(sheet_name)

        ws.append(["BARCODE", "ARTICLE", "НАЗВАНИЕ", "СКЛАД", "КОЛИЧЕСТВО"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
        for idx, width in enumerate([20, 16, 46, 34, 14], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"

        for row in rows:
            ws.append([
                row.get("barcode") or "", row.get("article", ""),
                row.get("name") or "", row.get("warehouse") or "",
                row.get("quantity", 0),
            ])

    if not wb.sheetnames:
        wb.create_sheet("Пусто")

    buffer = io.BytesIO()
    wb.save(buffer)

    name = _safe_filename(f"sklady_{store_slug}_{marketplace}")
    return buffer.getvalue(), f"{name}.xlsx"


def build_stock_xlsx(store_slug: str, store_name: str, marketplace: str,
                     columns: list[str], rows: list[list],
                     totals: list, ff_label: str = "") -> tuple[bytes, str]:
    """Основная таблица остатков магазина — как она сейчас на экране.

    columns/rows приходят готовыми: набор колонок зависит от площадки (у WB их
    две, у Ozon три, у Яндекса — по числу FBS-партнёров), и собирать его здесь
    заново значило бы описать эту логику во второй раз.

    Строка итогов идёт первой, сразу под шапкой, — так же, как в интерфейсе.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "для выгрузки в .xlsx нужен пакет openpyxl — установи его в .venv "
            "(pip install openpyxl)"
        ) from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _safe_filename(marketplace)[:31] or "Остатки"

    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.append(totals)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")

    for row in rows:
        ws.append(row)

    widths = [18, 18, 46] + [16] * (len(columns) - 3)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A3"

    # В имя файла кладём выбранный склад: иначе выгрузки по разным ФФ
    # ложатся в загрузки одинаковыми именами и различить их нельзя.
    parts = ["ostatki", store_slug, marketplace]
    if ff_label:
        parts.append(ff_label)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), _safe_filename("_".join(parts)) + ".xlsx"
