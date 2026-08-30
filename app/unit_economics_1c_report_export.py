import io
from datetime import date, datetime

_SUMMARY_COLUMNS = (
    ("name", "Товар", 38, "text"),
    ("article", "Артикул", 18, "text"),
    ("store_name", "Магазин", 14, "text"),
    ("subject", "Категория", 24, "text"),
    ("manager", "Менеджер", 22, "text"),
    ("funnel_period_from", "Период воронки с", 15, "date"),
    ("funnel_period_to", "Период воронки по", 15, "date"),
    ("orders_count", "Всего заказов, шт.", 17, "integer"),
    ("orders_amount", "Всего заказов, ₽", 18, "money"),
    ("cancel_count", "Всего отмен, шт.", 16, "integer"),
    ("cancel_amount", "Всего отмен, ₽", 17, "money"),
    ("net_orders_count", "Заказы − отмены, шт.", 20, "integer"),
    ("net_orders_amount", "ТО после отмен, ₽", 18, "money"),
    ("funnel_vendor_code", "Артикул продавца WB", 22, "text"),
    ("funnel_product_name", "Название товара WB", 34, "text"),
    ("funnel_source_version", "Версия дневных данных", 18, "integer"),
    ("funnel_updated_at", "Воронка обновлена", 22, "datetime"),
    ("buyout_count", "Выкупы, шт.", 16, "integer"),
    ("buyout_amount", "Сумма выкупов, ₽", 18, "money"),
    ("buyout_percent", "Процент выкупа", 16, "percent"),
    ("stock", "Остаток, шт.", 14, "integer"),
    ("impressions", "Показы", 14, "integer"),
    ("clicks", "Клики", 14, "integer"),
    ("ctr", "CTR", 12, "percent"),
    ("cpc", "CPC, ₽", 14, "money"),
    ("advertising_spend", "Расходы на рекламу, ₽", 19, "money"),
    ("orders_count", "Заказы для рекламы, шт.", 19, "integer"),
    ("expected_buyouts", "Расчётные выкупы, шт.", 19, "number"),
    ("advertising_per_unit", "Реклама на выкупленную ед., ₽", 24, "money"),
    ("drr", "ДРР с выкупом", 15, "percent"),
    ("retail_price", "Цена без СПП, ₽", 17, "money"),
    ("customer_price", "Расчётная клиентская цена, ₽", 23, "money"),
    ("customer_price_source", "Источник клиентской цены", 23, "text"),
    ("average_order_price", "Средняя цена заказа, ₽", 20, "money"),
    ("delivery_wb_rub", "Доставка WB, ₽", 16, "money"),
    ("return_cost_rub", "Возврат, ₽", 14, "money"),
    ("volume_l", "Объём, л", 13, "number"),
    ("acceptance_coefficient", "Коэффициент приёмки", 19, "number"),
    ("paid_acceptance_cost", "Платная приёмка, ₽", 18, "money"),
    ("logistics_buyout_percent", "Выкуп для логистики", 18, "percent"),
    ("delivery_with_returns", "Логистика с возвратами, ₽", 22, "money"),
    ("acquiring_percent", "Эквайринг, %", 15, "percent"),
    ("acquiring_value", "Эквайринг, ₽", 15, "money"),
    ("subject_commission_percent", "Комиссия предмета, %", 19, "percent"),
    ("wb_extra_tariff_percent", "Доп. тариф WB, %", 17, "percent"),
    ("commission_percent", "Комиссия WB итого, %", 19, "percent"),
    ("commission_value", "Комиссия WB, ₽", 17, "money"),
    ("storage_wb_rub", "Хранение в день, ₽", 17, "money"),
    ("storage_days", "Оборачиваемость, дней", 18, "integer"),
    ("storage_sum", "Хранение на единицу, ₽", 20, "money"),
    ("team_commission_percent", "Комиссия команды, %", 19, "percent"),
    ("team_commission_value", "Комиссия команды, ₽", 19, "money"),
    ("purchase_cost", "Закупка единицы, ₽", 18, "money"),
    ("fulfillment_cost", "Фулфилмент единицы, ₽", 20, "money"),
    ("tax_system", "Система налогообложения", 20, "text"),
    ("vat_percent", "НДС, %", 12, "percent"),
    ("vat_value", "НДС, ₽", 14, "money"),
    ("usn_percent", "УСН, %", 12, "percent"),
    ("usn_value", "УСН, ₽", 14, "money"),
    ("osno_percent", "ОСНО, %", 12, "percent"),
    ("osno_value", "ОСНО, ₽", 14, "money"),
    ("tax_value", "Налоги итого, ₽", 16, "money"),
    ("net_revenue", "Доход после удержаний, ₽", 20, "money"),
    ("margin_orders_count", "Выкупы в расчёте маржи", 21, "number"),
    ("margin", "Маржа периода, ₽", 18, "money"),
    ("purchase_value", "Закупка периода, ₽", 18, "money"),
    ("roi", "ROI", 13, "percent"),
)

_SUMMARY_GROUPS = (
    (1, 5, "Товар"),
    (6, 17, "Воронка за период отчёта"),
    (18, 20, "Выкуп за период отчёта"),
    (21, 21, "Остатки"),
    (22, 30, "Реклама"),
    (31, 34, "Цены"),
    (35, 41, "Логистика"),
    (42, 52, "Комиссии и хранение"),
    (53, 63, "Себестоимость и налоги"),
    (64, 67, "Результат"),
)

_DAILY_COLUMNS = (
    ("store_slug", "Магазин", 15, "text"),
    ("day", "Дата", 14, "date"),
    ("article", "nmId", 18, "text"),
    ("vendor_code", "Артикул продавца", 22, "text"),
    ("product_name", "Название WB", 38, "text"),
    ("orders_count", "Всего заказов, шт.", 18, "integer"),
    ("orders_amount", "Всего заказов, ₽", 18, "money"),
    ("cancel_count", "Всего отмен, шт.", 17, "integer"),
    ("cancel_amount", "Всего отмен, ₽", 17, "money"),
    ("buyout_count", "Выкупы, шт.", 16, "integer"),
    ("buyout_amount", "Сумма выкупов, ₽", 18, "money"),
    ("buyout_percent", "Процент выкупа", 17, "percent"),
    ("net_orders_count", "Заказы − отмены, шт.", 20, "integer"),
    ("net_orders_amount", "ТО после отмен, ₽", 18, "money"),
    ("source_version", "Версия данных", 15, "integer"),
    ("updated_at", "Обновлено", 22, "datetime"),
)

_WEEKLY_COLUMNS = (
    ("store_slug", "Магазин", 15, "text"),
    ("article", "nmId", 18, "text"),
    ("period_from", "Период с", 14, "date"),
    ("period_to", "Период по", 14, "date"),
    ("orders_count", "Всего заказов, шт.", 18, "integer"),
    ("orders_amount", "Всего заказов, ₽", 18, "money"),
    ("cancel_count", "Всего отмен, шт.", 17, "integer"),
    ("cancel_amount", "Всего отмен, ₽", 17, "money"),
    ("net_orders_count", "Заказы − отмены, шт.", 20, "integer"),
    ("net_orders_amount", "ТО после отмен, ₽", 18, "money"),
    ("buyout_percent", "Процент выкупа", 17, "percent"),
    ("source_version", "Версия данных", 15, "integer"),
    ("updated_at", "Обновлено", 22, "datetime"),
)

_DAILY_CALCULATION_COLUMNS = (
    ("advertising_spend", "Расходы на рекламу, ₽", 19, "money"),
    ("orders_count", "Заказы, шт.", 14, "integer"),
    ("net_orders_count", "Заказы − отмены, шт.", 20, "integer"),
    ("buyout_percent", "Выкуп, %", 14, "percent"),
    ("vat_percent", "НДС, %", 12, "percent"),
    ("usn_percent", "УСН, %", 12, "percent"),
    ("customer_price", "Цена с СПП, ₽", 17, "money"),
    ("retail_price", "Цена без СПП, ₽", 18, "money"),
    ("acquiring_percent", "Эквайринг, %", 15, "percent"),
    ("logistics", "Логистика, ₽", 15, "money"),
    ("storage", "Хранение, ₽", 15, "money"),
    ("commission_percent", "Комиссия WB, %", 18, "percent"),
    ("team_commission_percent", "Комиссия компании, %", 21, "percent"),
    ("fulfillment_cost", "Фулфилмент, ₽", 17, "money"),
    ("purchase_price", "Закупочная цена, ₽", 19, "money"),
    ("net_profit", "Чистая прибыль, ₽", 18, "money"),
    ("net_revenue", "Чистая выручка, ₽", 18, "money"),
    ("advertising_per_unit", "Реклама за 1 шт., ₽", 19, "money"),
    ("vat_value", "НДС, ₽", 14, "money"),
    ("usn_value", "УСН, ₽", 14, "money"),
)


def _number(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _date(value: object) -> date | None:
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _datetime(value: object) -> datetime | None:
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.replace(tzinfo=None)


def _cell_value(row: dict, key: str, kind: str) -> object:
    if key.startswith("daily:"):
        _, day_value, daily_key = key.split(":", 2)
        daily_row = next(
            (
                item
                for item in row.get("daily_calculations") or []
                if str(item.get("date")) == day_value
            ),
            {},
        )
        value = daily_row.get(daily_key)
        return _cell_value({daily_key: value}, daily_key, kind)
    if key == "expected_buyouts":
        return round(
            float(row.get("orders_count") or 0) * float(row.get("buyout_percent") or 0) / 100,
            2,
        )
    value = row.get(key)
    if kind == "text":
        return str(value or "")
    if kind == "integer":
        number = _number(value)
        return int(number) if number is not None else None
    if kind in {"number", "money"}:
        return _number(value)
    if kind == "percent":
        number = _number(value)
        return number / 100 if number is not None else None
    if kind == "date":
        return _date(value)
    if kind == "datetime":
        return _datetime(value)
    return value


def _number_format(kind: str) -> str:
    return {
        "integer": "#,##0",
        "number": "#,##0.00",
        "money": '#,##0.00" ₽"',
        "percent": "0.00%",
        "date": "dd.mm.yyyy",
        "datetime": "dd.mm.yyyy hh:mm",
    }.get(kind, "General")


def _style_header(sheet, columns, header_row: int) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    fill = PatternFill("solid", fgColor="E9E7FA")
    border = Border(bottom=Side(style="medium", color="6750D8"))
    for column, (_, title, width, _) in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=column, value=title)
        cell.fill = fill
        cell.font = Font(name="Arial", size=9, bold=True, color="18233F")
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = border
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.row_dimensions[header_row].height = 42


def _write_rows(sheet, rows: list[dict], columns, start_row: int) -> int:
    from openpyxl.styles import Alignment, Border, Font, Side

    border = Border(bottom=Side(style="thin", color="E4E7EC"))
    for row_index, row in enumerate(rows, start=start_row):
        for column_index, (key, _, _, kind) in enumerate(columns, start=1):
            cell = sheet.cell(
                row=row_index,
                column=column_index,
                value=_cell_value(row, key, kind),
            )
            cell.font = Font(name="Arial", size=9, color="18233F")
            cell.alignment = Alignment(
                vertical="top",
                horizontal="left" if kind == "text" else "right",
                wrap_text=kind == "text",
            )
            cell.border = border
            cell.number_format = _number_format(kind)
    return start_row + len(rows)


def _build_summary(workbook, report: dict) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    sheet = workbook.active
    sheet.title = "Расчёт маржи и ROI"
    sheet.sheet_view.showGridLines = False
    rows = (
        list(report.get("category_rows") or [])
        if report.get("group_by") == "subject"
        else list(report.get("rows") or [])
    )
    daily_dates = sorted(
        {
            str(item.get("date"))
            for row in rows
            for item in row.get("daily_calculations") or []
            if item.get("date")
        }
    )
    daily_columns = tuple(
        (f"daily:{day_value}:{key}", title, width, kind)
        for day_value in daily_dates
        for key, title, width, kind in _DAILY_CALCULATION_COLUMNS
    )
    columns = _SUMMARY_COLUMNS + daily_columns
    last_column = get_column_letter(len(columns))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = (
        "Отчёт по юниточной прибыли — по категориям"
        if report.get("group_by") == "subject"
        else "Отчёт по юниточной прибыли — полный расчёт"
    )
    sheet["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="18233F")
    sheet["A1"].alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 30
    sheet["A2"] = "Период"
    sheet["B2"] = _date(report.get("period_from"))
    sheet["C2"] = "—"
    sheet["D2"] = _date(report.get("period_to"))
    sheet["F2"] = "Категорий" if report.get("group_by") == "subject" else "Позиций"
    sheet["G2"] = len(rows)
    sheet["B2"].number_format = sheet["D2"].number_format = "dd.mm.yyyy"
    for cell in (sheet["A2"], sheet["F2"]):
        cell.font = Font(name="Arial", size=10, bold=True, color="5E6678")
    for cell in (sheet["B2"], sheet["D2"], sheet["G2"]):
        cell.font = Font(name="Arial", size=10, bold=True, color="18233F")
    sheet.merge_cells(f"A3:{last_column}3")
    sheet["A3"] = (
        "Источники: воронка продаж WB, рекламная статистика WB, остатки и параметры "
        "юнит-экономики 1С. Маржа периода = сумма дневных произведений маржи на штуку "
        "и заказов соответствующего дня; отмены показаны отдельно."
    )
    if not bool((report.get("totals") or {}).get("margin_complete", True)):
        missing = ", ".join((report.get("totals") or {}).get("margin_missing_days") or [])
        sheet["A3"] = f"{sheet['A3'].value} История маржи неполная{': ' + missing if missing else ''}."
    sheet["A3"].font = Font(name="Arial", size=9, italic=True, color="6F7788")
    sheet["A3"].alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    sheet.row_dimensions[3].height = 28

    for start, end, title in _SUMMARY_GROUPS:
        if end > start:
            sheet.merge_cells(start_row=4, start_column=start, end_row=4, end_column=end)
        for column in range(start, end + 1):
            sheet.cell(row=4, column=column).fill = PatternFill("solid", fgColor="6750D8")
        cell = sheet.cell(row=4, column=start, value=title)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", horizontal="center")
    daily_start = len(_SUMMARY_COLUMNS) + 1
    for day_index, day_value in enumerate(daily_dates):
        start = daily_start + day_index * len(_DAILY_CALCULATION_COLUMNS)
        end = start + len(_DAILY_CALCULATION_COLUMNS) - 1
        sheet.merge_cells(start_row=4, start_column=start, end_row=4, end_column=end)
        for column in range(start, end + 1):
            sheet.cell(row=4, column=column).fill = PatternFill(
                "solid",
                fgColor="5B7DB1" if day_index % 2 == 0 else "426A9F",
            )
        cell = sheet.cell(row=4, column=start, value=_date(day_value))
        cell.number_format = "dd.mm.yyyy"
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center", horizontal="center")
    _style_header(sheet, columns, 5)

    total_row = _write_rows(sheet, rows, columns, 6)
    totals = report.get("totals") or {}
    total_fill = PatternFill("solid", fgColor="E3DFFA")
    total_border = Border(top=Side(style="medium", color="6750D8"))
    total_values = {
        "name": "ИТОГО",
        "article": f"{len(rows)} {'кат.' if report.get('group_by') == 'subject' else 'поз.'}",
        "orders_count": totals.get("orders_count"),
        "orders_amount": totals.get("orders_amount"),
        "cancel_count": totals.get("cancel_count"),
        "cancel_amount": totals.get("cancel_amount"),
        "net_orders_count": totals.get("net_orders_count"),
        "net_orders_amount": totals.get("net_orders_amount"),
        "buyout_count": totals.get("buyout_count"),
        "buyout_amount": totals.get("buyout_amount"),
        "buyout_percent": totals.get("buyout_percent"),
        "stock": totals.get("stock"),
        "impressions": totals.get("impressions"),
        "clicks": totals.get("clicks"),
        "ctr": totals.get("ctr"),
        "cpc": totals.get("cpc"),
        "advertising_spend": totals.get("advertising_spend"),
        "drr": totals.get("drr"),
        "margin_orders_count": totals.get("margin_orders_count"),
        "margin": totals.get("margin"),
        "purchase_value": totals.get("purchase_value"),
        "roi": totals.get("roi"),
    }
    seen_keys: set[str] = set()
    for column_index, (key, _, _, kind) in enumerate(columns, start=1):
        value = total_values.get(key) if key not in seen_keys else None
        seen_keys.add(key)
        cell = sheet.cell(
            row=total_row,
            column=column_index,
            value=_cell_value({key: value}, key, kind) if value is not None else None,
        )
        cell.fill = total_fill
        cell.font = Font(name="Arial", size=10, bold=True, color="18233F")
        cell.alignment = Alignment(vertical="center", horizontal="left" if kind == "text" else "right")
        cell.border = total_border
        cell.number_format = _number_format(kind)
    sheet.row_dimensions[total_row].height = 24

    key_columns = {key: index for index, (key, _, _, _) in enumerate(columns, start=1)}
    for row_index in range(6, total_row):
        for key in ("margin", "roi"):
            value = _number(rows[row_index - 6].get(key))
            sheet.cell(row=row_index, column=key_columns[key]).font = Font(
                name="Arial",
                size=9,
                bold=True,
                color="168A5B" if (value or 0) >= 0 else "C74343",
            )

    data_end_row = max(total_row - 1, 5)
    sheet.auto_filter.ref = f"A5:{last_column}{data_end_row}"
    sheet.freeze_panes = "F6"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:5"
    sheet.print_area = f"A1:{last_column}{total_row}"


def _build_raw_sheet(workbook, title: str, rows: list[dict], columns, report: dict) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    last_column = get_column_letter(len(columns))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="18233F")
    sheet["A1"].alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 28
    sheet["A2"] = "Период отчёта"
    sheet["B2"] = _date(report.get("period_from"))
    sheet["C2"] = _date(report.get("period_to"))
    sheet["B2"].number_format = sheet["C2"].number_format = "dd.mm.yyyy"
    sheet.merge_cells(f"A3:{last_column}3")
    sheet["A3"] = "Исходные сохранённые показатели WB; разница рассчитана без изменения исходных заказов и отмен."
    sheet["A3"].font = Font(name="Arial", size=9, italic=True, color="6F7788")
    _style_header(sheet, columns, 4)
    end_row = _write_rows(sheet, rows, columns, 5)
    sheet.auto_filter.ref = f"A4:{last_column}{max(end_row - 1, 4)}"
    sheet.freeze_panes = "F5"
    sheet.print_title_rows = "1:4"


def build_xlsx(report: dict) -> tuple[bytes, str]:
    try:
        import openpyxl
    except ImportError as error:
        raise RuntimeError("для выгрузки в .xlsx нужен пакет openpyxl") from error

    workbook = openpyxl.Workbook()
    workbook.properties.creator = "CheckStock"
    workbook.properties.title = "Отчёт по юниточной прибыли"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    _build_summary(workbook, report)
    _build_raw_sheet(
        workbook,
        "Воронка по дням",
        list(report.get("funnel_daily_rows") or []),
        _DAILY_COLUMNS,
        report,
    )
    _build_raw_sheet(
        workbook,
        "Воронка 7 дней",
        list(report.get("funnel_weekly_rows") or []),
        _WEEKLY_COLUMNS,
        report,
    )
    buffer = io.BytesIO()
    workbook.save(buffer)
    filename = f"unit_profit_{report.get('period_from', '')}_{report.get('period_to', '')}.xlsx"
    return buffer.getvalue(), filename
