import io
from datetime import date


COLUMNS = (
    ("name", "Товар"), ("article", "Артикул"), ("store_name", "Магазин"),
    ("current_price", "Текущая цена с СПП и кошельком, ₽"),
    ("current_drr", "Текущий ДРР, %"), ("current_roi", "Текущий ROI, %"),
    ("target_price", "Целевая цена с СПП и кошельком, ₽"),
    ("target_drr", "Целевой ДРР, %"), ("target_roi", "Целевой ROI, %"),
)


def build_xlsx(rows: list[dict], period_from: str = "", period_to: str = "") -> tuple[bytes, str]:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Целевая цена"
    sheet.append([label for _, label in COLUMNS])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6757C8")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in rows:
        values = []
        for key, _ in COLUMNS:
            value = row.get(key)
            if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                value = "'" + value
            values.append(value)
        sheet.append(values)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{max(1, sheet.max_row)}"
    widths = (36, 22, 18, 22, 18, 18, 22, 18, 18)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=9):
        for cell in row:
            cell.number_format = '#,##0.00'
    buffer = io.BytesIO()
    workbook.save(buffer)
    suffix = f"_{period_from}_{period_to}" if period_from and period_to else f"_{date.today().isoformat()}"
    return buffer.getvalue(), f"target_price{suffix}.xlsx"
