import io

from app import db
from app.formatting import format_dt
from app.stock_cost_report import fbs_sales_for_view, operations_for_view
from app.stores import STORES


def _safe_filename(value: str) -> str:
    cleaned = "".join(character for character in value if character.isalnum() or character in " _-.()")
    return cleaned.strip() or "stock_cost_report"


def build_xlsx(report: dict, view: str) -> tuple[bytes, str]:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError("для выгрузки в .xlsx нужен пакет openpyxl") from error

    workbook = openpyxl.Workbook()
    operations_sheet = workbook.active
    operations_sheet.title = "Операции"
    items_sheet = workbook.create_sheet("По артикулам")

    def prepare(sheet, titles: list[str], widths: list[int]) -> None:
        sheet.append(titles)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(titles))}1"

    prepare(
        operations_sheet,
        [
            "Когда / период",
            "Магазин",
            "Маркетплейс",
            "Операция",
            "Откуда",
            "Куда",
            "Позиций",
            "Количество",
            "ЗЦ, руб.",
            "Без ЗЦ, ед.",
            "Источник",
            "Примечание",
            "Сотрудник",
        ],
        [22, 16, 18, 24, 30, 30, 11, 13, 16, 13, 18, 34, 24],
    )
    prepare(
        items_sheet,
        [
            "Когда / период",
            "Магазин",
            "Маркетплейс",
            "Операция",
            "Откуда",
            "Куда",
            "BARCODE",
            "ARTICLE",
            "Название",
            "Количество",
            "ЗЦ за ед., руб.",
            "ЗЦ всего, руб.",
            "FBS на начало",
            "Перемещено на FBS",
            "FBS на конец",
            "Продажи по формуле",
            "Примечание",
            "Сотрудник",
        ],
        [22, 16, 18, 24, 30, 30, 20, 18, 46, 13, 16, 16, 14, 18, 14, 19, 34, 24],
    )

    operations = operations_for_view(report, view)
    for operation in reversed(operations):
        marketplace = operation.get("to_marketplace") or operation.get("from_marketplace") or ""
        operation_row = [
            format_dt(operation.get("created_at")),
            (
                STORES[operation["store_slug"]]["name"]
                if operation["store_slug"] in STORES
                else operation["store_slug"].upper()
            ),
            marketplace,
            operation["label"],
            operation["from_label"],
            operation["to_label"],
            operation["positions"],
            operation["units"],
            operation["purchase_cost"],
            operation["missing_units"],
            db.SOURCE_LABELS.get(operation.get("source_type"), operation.get("source_type") or ""),
            operation.get("note") or "",
            operation.get("user_name") or "",
        ]
        operations_sheet.append(operation_row)
        for item in operation["items"]:
            items_sheet.append(
                [
                    operation_row[0],
                    operation_row[1],
                    marketplace,
                    operation["label"],
                    operation["from_label"],
                    operation["to_label"],
                    item.get("barcode") or "",
                    item.get("article") or "",
                    item.get("name") or "",
                    item.get("quantity") or 0,
                    item.get("purchase_price"),
                    item.get("purchase_cost"),
                    "",
                    "",
                    "",
                    "",
                    operation.get("note") or "",
                    operation.get("user_name") or "",
                ]
            )

    sales = fbs_sales_for_view(report, view)
    sales_by_key: dict[tuple[str, str], list[dict]] = {}
    for item in sales:
        sales_by_key.setdefault((item["store_slug"], item["marketplace"]), []).append(item)
    reconciliation_items = {
        (item["store_slug"], item["marketplace"], article["article"]): article
        for item in report["reconciliation"]
        if item["available"]
        for article in item["items"]
    }
    period_label = f"{report['date_from'].isoformat()} — {report['date_to'].isoformat()}"
    for (store_slug, marketplace), group in sorted(sales_by_key.items()):
        cost = round(
            sum(float(item["purchase_cost"]) for item in group if item["purchase_cost"] is not None),
            2,
        )
        missing_units = sum(abs(int(item["quantity"])) for item in group if item["purchase_cost"] is None)
        store_name = STORES[store_slug]["name"] if store_slug in STORES else store_slug.upper()
        operations_sheet.append(
            [
                period_label,
                store_name,
                marketplace,
                "Продажи FBS",
                "Сток FBS",
                "Покупатели",
                len(group),
                sum(int(item["quantity"]) for item in group),
                cost,
                missing_units,
                "Данные продаж",
                "Выкупленные товары FBS/rFBS",
                "Система",
            ]
        )
        for item in group:
            formula = reconciliation_items.get((store_slug, marketplace, item["article"]), {})
            items_sheet.append(
                [
                    period_label,
                    store_name,
                    marketplace,
                    "Продажи FBS",
                    "Сток FBS",
                    "Покупатели",
                    item.get("barcode") or "",
                    item.get("article") or "",
                    item.get("name") or "",
                    item.get("quantity") or 0,
                    item.get("purchase_price"),
                    item.get("purchase_cost"),
                    formula.get("start_quantity", ""),
                    formula.get("moved_quantity", ""),
                    formula.get("end_quantity", ""),
                    formula.get("quantity", ""),
                    "Выкупленные товары FBS/rFBS",
                    "Система",
                ]
            )

    for sheet in (operations_sheet, items_sheet):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in range(2, operations_sheet.max_row + 1):
        operations_sheet.cell(row=row, column=9).number_format = '#,##0.00" ₽"'
    for row in range(2, items_sheet.max_row + 1):
        for column in (11, 12):
            items_sheet.cell(row=row, column=column).number_format = '#,##0.00" ₽"'

    buffer = io.BytesIO()
    workbook.save(buffer)
    filename = _safe_filename(
        f"dvizhenie_zc_{report['date_from'].isoformat()}_{report['date_to'].isoformat()}"
    )
    return buffer.getvalue(), f"{filename}.xlsx"
