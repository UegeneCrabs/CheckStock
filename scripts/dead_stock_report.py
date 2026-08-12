import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app import db
from app.stores import STORES
from app.wb import api as wb_api
from app.wb import tokens as wb_tokens

MARKETPLACE = "WB"


def _store_label(store_slug: str) -> str:
    return STORES.get(store_slug, {}).get("name") or store_slug.upper()


def count_orders(orders: list[dict], since: datetime) -> tuple[dict, dict, int]:

    by_barcode: dict[str, int] = {}
    by_nm: dict[str, int] = {}
    cancelled = 0

    for order in orders:
        raw_date = str(order.get("date") or "")[:19]
        try:
            ordered_at = datetime.fromisoformat(raw_date)
        except ValueError:
            continue
        if ordered_at < since:
            continue

        if order.get("isCancel"):
            cancelled += 1
            continue

        barcode = str(order.get("barcode") or "").strip()
        nm_id = str(order.get("nmId") or "").strip()
        if barcode:
            by_barcode[barcode] = by_barcode.get(barcode, 0) + 1
        if nm_id:
            by_nm[nm_id] = by_nm.get(nm_id, 0) + 1

    return by_barcode, by_nm, cancelled


def orders_for_item(item: dict, by_barcode: dict, by_nm: dict) -> int:

    barcode = str(item.get("barcode") or "").strip()
    if barcode in by_barcode:
        return by_barcode[barcode]

    article = str(item["article"])
    if " / " in article:
        return 0
    return by_nm.get(article, 0)


def analyze_store(store_slug: str, weeks: int) -> dict:

    label = _store_label(store_slug)
    since = datetime.now() - timedelta(weeks=weeks)

    catalog = db.get_stock_items(store_slug, MARKETPLACE, ("fbs", "fbo"))
    if not catalog:
        return {"store": store_slug, "label": label, "error": "каталог пуст — сначала синхронизация"}

    try:
        orders = wb_api.get_orders(wb_tokens.get_token(store_slug), since.strftime("%Y-%m-%dT%H:%M:%S"))
    except Exception as e:
        return {"store": store_slug, "label": label, "error": str(e)}

    by_barcode, by_nm, cancelled = count_orders(orders, since)

    dead = []
    for item in catalog:
        fbs = item.get("fbs_stock") or 0
        fbo = item.get("fbo_stock") or 0
        sold = orders_for_item(item, by_barcode, by_nm)

        if sold or fbs or fbo:
            continue

        dead.append(
            {
                "article": item["article"],
                "barcode": item["barcode"],
                "name": item["name"],
                "ff": item.get("ff_available") or 0,
                "mp_updated_at": (item.get("mp_updated_at") or "")[:10],
            }
        )

    return {
        "store": store_slug,
        "label": label,
        "catalog": len(catalog),
        "orders": sum(by_barcode.values()) or sum(by_nm.values()),
        "cancelled": cancelled,
        "dead": dead,
    }


def build_xlsx(results: list[dict], weeks: int, path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise SystemExit("нужен пакет openpyxl: pip install openpyxl") from e

    book = openpyxl.Workbook()
    head_font = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDEBF7")

    def style_head(sheet) -> None:
        for cell in sheet[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        sheet.freeze_panes = "A2"

    summary = book.active
    summary.title = "Сводка"
    summary.append(
        ["МАГАЗИН", "ПОЗИЦИЙ В КАТАЛОГЕ", f"ЗАКАЗОВ ЗА {weeks} НЕД.", "БЕЗ ДВИЖЕНИЯ", "ДОЛЯ", "ПРИМЕЧАНИЕ"]
    )

    for res in results:
        if res.get("error"):
            summary.append([res["label"], "", "", "", "", res["error"]])
            continue
        share = len(res["dead"]) / res["catalog"] if res["catalog"] else 0
        summary.append([res["label"], res["catalog"], res["orders"], len(res["dead"]), round(share, 3), ""])

    for column, width in zip("ABCDEF", (24, 20, 20, 16, 10, 52), strict=True):
        summary.column_dimensions[column].width = width
    style_head(summary)

    for res in results:
        if res.get("error"):
            continue

        title = "".join(c for c in res["label"] if c not in "[]:*?/\\")[:31]
        sheet = book.create_sheet(title or res["store"][:31])
        sheet.append(["АРТИКУЛ", "ШТРИХКОД", "НАЗВАНИЕ", "ЗАКАЗОВ", "FBS", "FBO", "НА ФФ", "ИЗМЕНЕНА В ЛК"])
        for row in res["dead"]:
            sheet.append(
                [row["article"], row["barcode"], row["name"], 0, 0, 0, row["ff"], row["mp_updated_at"]]
            )
        for column, width in zip("ABCDEFGH", (18, 18, 46, 10, 8, 8, 10, 16), strict=True):
            sheet.column_dimensions[column].width = width
        style_head(sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Товары WB без заказов и без стока на площадке")
    parser.add_argument("--store", help="один магазин (slug), по умолчанию все")
    parser.add_argument("--weeks", type=int, default=3, help="период в неделях (по умолчанию 3)")
    parser.add_argument("--out", help="куда сохранить .xlsx")
    args = parser.parse_args()

    slugs = [args.store] if args.store else list(STORES)
    slugs = [slug for slug in slugs if wb_tokens.has_token(slug)]
    if not slugs:
        raise SystemExit("Нет магазинов с токеном WB")

    results = []
    for slug in slugs:
        print(f"{_store_label(slug)}: считаю...", flush=True)
        res = analyze_store(slug, args.weeks)
        results.append(res)

        if res.get("error"):
            print(f"  пропущен: {res['error']}")
            continue

        share = len(res["dead"]) / res["catalog"] if res["catalog"] else 0
        print(
            f"  каталог {res['catalog']}, заказов {res['orders']}"
            f" (отменённых {res['cancelled']}), без движения {len(res['dead'])}"
            f" — {share:.0%}"
        )

    out = (
        Path(args.out)
        if args.out
        else Path(f"data/exports/без-движения-{args.weeks}нед-{datetime.now():%Y-%m-%d}.xlsx")
    )
    build_xlsx(results, args.weeks, out)
    print(f"\nОтчёт: {out.resolve()}")


if __name__ == "__main__":
    main()
