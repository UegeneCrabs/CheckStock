import io
from pathlib import Path

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app import db, stock_total
from app.container import ApplicationContainer
from app.dto.identity import User
from app.stores import STORES
from app.web import middleware

NOW = "2026-08-21T12:00:00+03:00"
FULFILLMENT = "ФулСервис Подольск"


def _catalog_item(article: str, barcode: str, name: str) -> dict:
    return {"article": article, "barcode": barcode, "name": name}


def test_total_stock_merges_marketplaces_by_barcode_and_keeps_stores_separate(database_path) -> None:
    del database_path
    db.replace_catalog(
        "rimili",
        "WB",
        [_catalog_item("WB-ARTICLE", "2200000000001", "Общий товар")],
        NOW,
    )
    db.replace_catalog(
        "rimili",
        "OZON",
        [
            _catalog_item("OZON-ARTICLE", "2200000000001", "Общий товар Ozon"),
            _catalog_item("OZON-ZERO", "", "Товар без штрихкода"),
        ],
        NOW,
    )
    db.replace_catalog(
        "rimili",
        "YANDEX MARKET",
        [_catalog_item("YANDEX-ARTICLE", "2200000000001", "Общий товар Яндекс")],
        NOW,
    )
    db.replace_catalog(
        "tris",
        "WB",
        [_catalog_item("TRIS-ARTICLE", "2200000000001", "Такой же штрихкод, другой магазин")],
        NOW,
    )
    connection = db.get_connection()
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS catalog_product_exclusions (
            store_slug TEXT NOT NULL,
            marketplace TEXT NOT NULL,
            nm_id TEXT NOT NULL,
            status TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (store_slug, marketplace, nm_id)
        )
        """
    )
    connection.execute(
        """
        INSERT INTO catalog_product_exclusions
            (store_slug, marketplace, nm_id, status, updated_at)
        VALUES ('rimili', 'WB', 'WB-ARTICLE', 'Старье', ?)
        """,
        (NOW,),
    )
    wb_stock_item_id = connection.execute(
        """
        SELECT id FROM stock_items
         WHERE store_slug='rimili' AND marketplace='WB' AND article='WB-ARTICLE'
        """
    ).fetchone()[0]
    connection.execute(
        """
        INSERT INTO unit_economics_1c_source_values
            (stock_item_id, purchase_price, source_sheet_id, source_sheet_title,
             source_row, synced_at)
        VALUES (?, 100, 1, 'RIMILI WB', 2, ?)
        """,
        (wb_stock_item_id, NOW),
    )
    connection.commit()
    connection.close()

    db.upsert_ff_stock("rimili", "WB-ARTICLE", FULFILLMENT, 2, NOW, "WB")
    db.upsert_ff_stock("rimili", "OZON-ARTICLE", FULFILLMENT, 4, NOW, "OZON")
    db.upsert_mp_stock("rimili", "WB-ARTICLE", "WB", "fbs", 7, NOW)
    db.upsert_mp_stock("rimili", "OZON-ARTICLE", "OZON", "rfbs", 3, NOW)
    db.upsert_mp_stock("rimili", "YANDEX-ARTICLE", "YANDEX MARKET", "fbo", 5, NOW)
    db.upsert_mp_stock("tris", "TRIS-ARTICLE", "WB", "fbo", 11, NOW)
    connection = db.get_connection()
    cursor = connection.execute(
        """
        INSERT INTO ff_transit_batches
            (store_slug, from_fulfillment, from_marketplace, to_fulfillment,
             to_marketplace, status, note, sent_by_name, sent_at)
        VALUES ('rimili', 'Source', 'WB', 'Target', 'OZON',
                'in_transit', 'Total test', 'Tester', ?)
        RETURNING id
        """,
        (NOW,),
    )
    connection.execute(
        """
        INSERT INTO ff_transit_items
            (batch_id, from_article, to_article, barcode, name,
             sent_quantity, received_quantity, cancelled_quantity)
        VALUES (?, 'WB-ARTICLE', 'OZON-ARTICLE', '2200000000001',
                'Общий товар', 3, 0, 0)
        """,
        (cursor.lastrowid,),
    )
    connection.commit()
    connection.close()

    rows = stock_total.build_rows(("rimili", "tris"))
    shared = next(row for row in rows if row["store_slug"] == "rimili" and row["barcode"] == "2200000000001")
    tris = next(row for row in rows if row["store_slug"] == "tris" and row["barcode"] == "2200000000001")
    zero = next(row for row in rows if row["article"] == "OZON-ZERO")

    assert shared["article"] == "WB-ARTICLE"
    assert shared["ff_wb"] == 2
    assert shared["ff_ozon"] == 4
    assert shared["transit_ozon"] == 3
    assert shared["fbs_wb"] == 7
    assert shared["rfbs_ozon"] == 3
    assert shared["fbo_yandex"] == 5
    assert shared["total_wb"] == 9
    assert shared["total_ozon"] == 10
    assert shared["total_yandex"] == 5
    assert shared["grand_total"] == 24
    assert shared["purchase_price"] == 100
    assert tris["grand_total"] == 11
    assert zero["grand_total"] == 0
    assert all(row["grand_total"] >= rows[index + 1]["grand_total"] for index, row in enumerate(rows[:-1]))
    # Both total views load today's source value, not yesterday's margin snapshot.
    with db.get_connection() as connection:
        connection.execute(
            "UPDATE unit_economics_1c_source_values SET purchase_price=125.5 WHERE stock_item_id=?",
            (wb_stock_item_id,),
        )
        connection.commit()
    refreshed = stock_total.build_rows(("rimili",))
    updated = next(row for row in refreshed if row["article"] == "WB-ARTICLE")
    assert updated["purchase_price"] == 125.5
    from app.web.routers.stock_total import _render_rows, _render_totals

    rendered = _render_rows([updated])
    assert '<td data-filter-value="125.5">125,50 ₽</td>' in rendered
    assert rendered.count("<td") == 23
    assert '<th data-cost-total-column="4">3 012,00 ₽</th>' in _render_totals([updated])
    assert _render_totals([updated]).count("<th") == 46


def test_total_stock_xlsx_has_store_column_grouped_headers_and_zeroes(database_path) -> None:
    del database_path
    rows = [
        {
            "store_slug": "rimili",
            "store_name": "RIMILI",
            "article": "ARTICLE-1",
            "barcode": "0012345678901",
            "name": "Тестовый товар",
            "purchase_price": 100,
            "grand_total": 9,
            "total_wb": 9,
            "total_ozon": 0,
            "total_yandex": 0,
            **{key: (9 if key == "fbs_wb" else 0) for key in stock_total.QUANTITY_KEYS},
        }
    ]

    content, filename = stock_total.build_xlsx(rows)
    workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    sheet = workbook["Остатки Тотал"]

    assert filename.startswith("ostatki_total_")
    assert sheet["A1"].value == "МАГАЗИН"
    assert sheet["E1"].value == "ТЕКУЩАЯ ЗЦ, ₽"
    assert sheet["E5"].value == 100
    assert sheet["E5"].number_format == '#,##0.00 "₽"'
    assert sheet["E3"].value is None
    assert sheet["E4"].value is None
    assert sheet["F1"].value == "ТОТАЛ"
    assert sheet["F2"].value == "ГРАНД ТОТАЛ"
    assert sheet["G2"].value == "ВБ"
    assert sheet["H2"].value == "ОЗОН"
    assert sheet["I2"].value == "ЯМ"
    assert sheet["J1"].value == "ДОСТУПНО ФФ ДЛЯ РАСПРЕДЕЛЕНИЯ"
    assert sheet["M1"].value == "В ПУТИ МЕЖДУ ФФ"
    assert sheet["P1"].value == "ТЕКУЩИЙ СТОК В ПРОДАЖЕ FBS"
    assert sheet["S1"].value == "ТЕКУЩИЙ СТОК В ПРОДАЖЕ RFBS"
    assert sheet["V1"].value == "ТЕКУЩИЙ СТОК В ПРОДАЖЕ FBO"
    assert {str(item) for item in sheet.merged_cells.ranges} >= {
        "A1:A2",
        "F1:I1",
        "J1:L1",
        "M1:O1",
        "P1:R1",
        "S1:U1",
        "V1:X1",
    }
    assert sheet["A3"].value == "ИТОГО"
    assert sheet["A4"].value == "ИТОГО В ЗЦ"
    assert sheet["F4"].value == 900
    assert sheet["A5"].value == "RIMILI"
    assert sheet["B5"].value == "ARTICLE-1"
    assert sheet["C5"].value == "0012345678901"
    assert sheet["F5"].value == 9
    assert sheet["G5"].value == 9
    assert sheet["J5"].value == 0
    assert sheet["M5"].value == 0
    assert sheet["P5"].value == 9
    assert sheet.freeze_panes == "J5"


def test_total_routes_resolve_before_the_dynamic_store_route(
    application: FastAPI,
    client: TestClient,
    user_factory,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    user: User = user_factory()
    container: ApplicationContainer = application.state.container
    monkeypatch.setattr(container.identity, "user_for_token", lambda _token: user)
    client.cookies.set(middleware.auth.SESSION_COOKIE, "stock-total-test")

    page = client.get("/stock/total")
    download = client.get("/stock/total.xlsx")
    store_page = client.get("/stock/rimili", params={"mp": "WB"})
    filtered_page = client.get("/stock/total", params={"store": "rimili"})
    filtered_download = client.get("/stock/total.xlsx", params={"store": "rimili"})

    assert page.status_code == 200
    assert "ТЕКУЩАЯ ЗЦ, ₽" in page.text
    assert "ТЕКУЩАЯ ЗЦ, ₽" in store_page.text
    assert 'data-filter-column="22"' in page.text
    assert "Остатки Тотал" in page.text
    assert 'id="stock-total-table"' in page.text
    assert 'class="topbar"' not in page.text
    assert 'class="stock-total-summary"' not in page.text
    controls = page.text.split('<div class="stock-total-controls">', 1)[1].split("</div>", 1)[0]
    assert 'class="btn-primary stock-total-download"' in controls
    assert controls.index("FBO") < controls.index("Скачать XLSX")
    assert store_page.status_code == 200
    assert 'class="mp-tab mp-tab--total"' in store_page.text
    assert 'data-store-total>ТОТАЛ</button>' in store_page.text
    assert 'href="/stock/total?store=rimili">ТОТАЛ</a>' not in store_page.text
    assert 'id="store-stock-total-table"' in store_page.text
    assert store_page.text.count('id="store-stock-total-table"') == 1
    assert 'data-stock-store-total-view hidden' in store_page.text
    assert "Тотал по трём площадкам" not in store_page.text
    for store_slug in STORES:
        cabinet_page = client.get(f"/stock/{store_slug}", params={"mp": "WB"})
        assert cabinet_page.status_code == 200
        assert cabinet_page.text.count('id="store-stock-total-table"') == 1
        assert "Тотал по трём площадкам" not in cabinet_page.text
        assert 'data-stock-store-total-view hidden' in cabinet_page.text
    assert 'href="/stock/total.xlsx?store=rimili"' in store_page.text
    store_total_data = client.get("/stock/rimili/total-data")
    assert store_total_data.status_code == 200
    assert store_total_data.json()["store"] == "rimili"
    assert tuple(store_total_data.json()["total_keys"]) == stock_total.TOTAL_KEYS
    assert tuple(store_total_data.json()["quantity_keys"]) == stock_total.QUANTITY_KEYS
    assert filtered_page.status_code == 200
    assert '<option value="rimili" selected>RIMILI</option>' in filtered_page.text
    assert 'href="/stock/total.xlsx?store=rimili"' in filtered_page.text
    assert client.get("/stock/total", params={"store": "unknown"}).status_code == 404

    table_filter_script = (
        Path(__file__).resolve().parents[2] / "static" / "table-filter.js"
    ).read_text(encoding="utf-8")
    assert "if (th.querySelector('.tf-th-inner')) return;" in table_filter_script
    assert download.status_code == 200
    assert download.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert filtered_download.status_code == 200
    filtered_workbook = openpyxl.load_workbook(io.BytesIO(filtered_download.content), data_only=True)
    filtered_sheet = filtered_workbook["Остатки Тотал"]
    exported_stores = {
        filtered_sheet.cell(row=row_number, column=1).value
        for row_number in range(5, filtered_sheet.max_row + 1)
        if filtered_sheet.cell(row=row_number, column=1).value
    }
    assert exported_stores <= {"RIMILI"}

    stock_total_script = (
        Path(__file__).resolve().parents[2] / "static" / "stock-total.js"
    ).read_text(encoding="utf-8")
    assert "querySelectorAll('[data-total-column]')" in stock_total_script
    assert "querySelectorAll('[data-cost-total-column]')" in stock_total_script

    store_total_script = (
        Path(__file__).resolve().parents[2] / "static" / "store-total.js"
    ).read_text(encoding="utf-8")
    assert "'/total-data'" in store_total_script
    assert "data-store-cost-key" in store_total_script
    assert "button.addEventListener('click', loadTotal)" in store_total_script
    assert "window.location =" not in store_total_script

    total_styles = (Path(__file__).resolve().parents[2] / "static" / "stock-total.css").read_text(
        encoding="utf-8"
    )
    assert "[data-stock-store-total-view][hidden]" in total_styles
    assert "width: calc(47% / 15)" in total_styles
    assert "scrollbar-gutter: auto" in total_styles
