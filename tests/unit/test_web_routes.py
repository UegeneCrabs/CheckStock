import json
import logging
import tempfile
import time
import unittest
from datetime import UTC, date, datetime
from pathlib import Path
from threading import Event
from unittest import mock

from fastapi.testclient import TestClient

from app import db, decision_center, rnp_analytics
from app.dto.decision import DecisionAction, DecisionStatus
from app.dto.marketplace import Marketplace
from app.dto.rnp import RnpAction, RnpStrategy
from app.dto.system import ReadinessStatus
from app.dto.unit_economics_1c import UnitEconomics1CProductSettings
from app.main import create_app
from app.repositories import core
from app.stores import STORES
from app.wb import funnel_orders as wb_funnel_orders
from app.web import middleware
from app.web.identifiers import copy_identifier
from app.web.routers import decision_center as decision_routes
from app.web.routers import rnp as rnp_routes
from app.web.routers import sales_overview, unit_economics

NOW = "2026-08-12T10:00:00+00:00"


class WebRouteUnitTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.database_patch = mock.patch.object(
            core,
            "DB_PATH",
            Path(self.temporary_directory.name) / "web-unit.db",
        )
        self.database_patch.start()
        db.init_db()
        decision_center.init_schema()
        rnp_analytics.init_schema()
        db.seed_defaults()
        self.user = {
            "id": 1,
            "full_name": "Unit Admin",
            "login": "unit-admin",
            "role": "superadmin",
            "is_active": 1,
            "can_edit_stock": 1,
            "can_manage_users": 1,
            "store_slugs": list(STORES),
        }
        self.app = create_app()
        self.authentication_patch = mock.patch.object(
            self.app.state.container.identity,
            "user_for_token",
            return_value=self.user,
        )
        self.authentication_patch.start()
        logging.disable(logging.CRITICAL)
        self.client = TestClient(self.app, raise_server_exceptions=False)
        self.client.cookies.set(middleware.auth.SESSION_COOKIE, "test-session")
        db.upsert_ff_stock("rimili", "949558341", "ФулСервис Подольск", 5, NOW, "WB")
        db.upsert_mp_stock("rimili", "949558341", "WB", "fbs", 3, NOW)
        db.replace_mp_warehouse_stock(
            "rimili",
            "WB",
            "fbo",
            [("949558341", "Коледино", "Москва", 2, NOW)],
        )
        self.operation_id = db.record_operation(
            "rimili",
            "delivery",
            "manual",
            [
                {
                    "article": "949558341",
                    "barcode": "2050292584830",
                    "name": "Ретро гирлянда",
                    "quantity": 5,
                }
            ],
            1,
            "Unit Admin",
            NOW,
        )

    def _unit_economics_products(self) -> list[dict]:
        response = self.client.get("/sales/unit-economics-1c", params={"data": "1"})
        self.assertEqual(response.status_code, 200, response.text)
        self.assertTrue(response.json()["ok"])
        return response.json()["products"]

    def _unit_economics_product(self, store_slug: str, article: str) -> dict:
        response = self.client.get(
            "/sales/unit-economics-1c",
            params={"data": "1", "store": store_slug, "article": article},
        )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertTrue(response.json()["ok"])
        return response.json()["product"]

    def tearDown(self) -> None:
        self.client.close()
        logging.disable(logging.NOTSET)
        self.authentication_patch.stop()
        self.database_patch.stop()
        self.temporary_directory.cleanup()

    def test_html_pages_and_downloads_render(self) -> None:
        paths = (
            "/sales",
            "/sales/decision-center",
            "/sales/ephemerides",
            "/sales/rnp",
            "/sales/unit-economics-1c/cabinet-settings",
            "/sales/unit-economics-1c",
            "/sales/unit-economics-1c/reports/unit-profit",
            "/sales/unit-economics-1c/ozon",
            "/sales/unit-economics-1c/yandex-market",
            "/supply",
            "/stock",
            "/stock/total",
            "/stock/supplies",
            "/stock/randomizer",
            "/stock-2",
            "/stock-2/details/frozen",
            "/stock/rimili",
            "/stock/rimili/fbs",
            "/stock/rimili/warehouses",
            "/stock/rimili/operations",
            "/stock/cost-report",
            "/admin",
            "/admin/activity",
            "/admin/google-export",
            "/admin/integrations",
        )
        for path in paths:
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertEqual(response.status_code, 200, f"{path}: {response.text[:500]}")
                self.assertTrue(response.content)
                if "text/html" in response.headers.get("content-type", ""):
                    self.assertIn('<header class="sidebar app-header" id="app-header">', response.text)
                    self.assertNotIn('<aside class="sidebar"', response.text)
                    self.assertNotIn('class="topbar"', response.text)
                    self.assertNotIn("Система работает", response.text)

        for path in (
            "/sales",
            "/sales/ephemerides",
            "/sales/rnp",
            "/stock",
            "/stock/total",
            "/stock/supplies",
            "/stock/randomizer",
            "/stock-2",
            "/stock-2/details/frozen",
            "/stock/rimili",
            "/stock/cost-report",
            "/sales/unit-economics-1c/cabinet-settings",
            "/sales/unit-economics-1c",
            "/sales/unit-economics-1c/reports/unit-profit",
            "/sales/unit-economics-1c/ozon",
            "/sales/unit-economics-1c/yandex-market",
        ):
            with self.subTest(data_workspace_path=path):
                response = self.client.get(path)
                self.assertNotIn('class="topbar"', response.text)
                self.assertNotIn("Система работает", response.text)

        stock_total_page = self.client.get("/stock/total")
        self.assertNotIn("ЕДИНЫЙ СРЕЗ ПО ВСЕМ КАБИНЕТАМ", stock_total_page.text)
        self.assertNotIn("Позиции разных площадок объединены по штрихкоду", stock_total_page.text)
        sales_page = self.client.get("/sales")
        self.assertNotIn("Продажи по магазинам и схемам", sales_page.text)

        unit_page = self.client.get("/sales/unit-economics-1c")
        self.assertNotIn("ЮНИТ-ЭКОНОМИКА 1С / WILDBERRIES", unit_page.text)
        self.assertNotIn("Юнит-экономика 1С / Wildberries", unit_page.text)

        root = self.client.get("/", follow_redirects=False)
        self.assertEqual(root.status_code, 303)
        self.assertEqual(root.headers["location"], "/sales")

        stock_page = self.client.get("/stock")
        self.assertIn('id="sync-products-btn"', stock_page.text)
        self.assertIn("Обновить остатки", stock_page.text)
        self.assertIn("catalogBadge(entry.wb_catalog)", stock_page.text)
        self.assertIn("if (!response.ok)", stock_page.text)
        self.assertIn('aria-live="polite"', stock_page.text)

        article = self.client.get(
            "/stock/rimili/article-detail",
            params={"mp": "WB", "article": "949558341"},
        )
        self.assertEqual(article.status_code, 200, article.text)
        self.assertIn("949558341", article.text)

        downloads = (
            "/stock/rimili/stock.xlsx?mp=WB",
            "/stock/rimili/warehouses/xlsx?mp=WB&scheme=fbo&warehouse=Коледино",
            "/stock/rimili/operations/xlsx",
            "/stock/cost-report.xlsx",
            f"/admin/operations/{self.operation_id}/xlsx",
        )
        for path in downloads:
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertEqual(response.status_code, 200, response.text[:500])
                self.assertTrue(response.content)

    def test_only_unit_economics_1c_is_visible_in_sales_navigation(self) -> None:
        response = self.client.get("/sales")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.text.count('class="nav-item nav-group-trigger'), 4)
        self.assertIn('aria-controls="nav-reports-submenu" data-nav-toggle', response.text)
        self.assertIn('id="nav-reports-submenu" data-nav-submenu', response.text)
        self.assertNotIn(
            'href="/sales/unit-economics-1c/reports/unit-profit" title=',
            response.text,
        )
        self.assertIn('href="/sales/unit-economics-1c"', response.text)
        self.assertNotIn('href="/sales/unit-economics-1c/cabinet-settings"', response.text)
        self.assertIn('title="Отчёты"', response.text)
        self.assertIn(">Юниточная прибыль</a>", response.text)
        self.assertNotIn('href="/sales/unit-economics/wb-fbs"', response.text)

    def test_product_identifiers_are_copyable_globally(self) -> None:
        root = Path(__file__).resolve().parents[2]
        page_template = (root / "templates" / "page.html").read_text(encoding="utf-8")
        copy_script = (root / "static" / "identifier-copy.js").read_text(encoding="utf-8")
        self.assertIn('/static/identifier-copy.css?v=1', page_template)
        self.assertIn('/static/identifier-copy.js?v=1', page_template)
        self.assertIn("navigator.clipboard.writeText(value)", copy_script)
        self.assertIn("document.execCommand('copy')", copy_script)
        self.assertIn("document.addEventListener('pointerover'", copy_script)
        self.assertIn("document.addEventListener('focusin'", copy_script)
        self.assertIn("event.stopPropagation()", copy_script)

        total_page = self.client.get("/stock/total")
        self.assertIn('data-copy-kind="Артикул"', total_page.text)
        self.assertIn('data-copy-kind="Баркод"', total_page.text)
        stock_page = self.client.get("/stock/rimili")
        self.assertIn('data-copy-value="949558341"', stock_page.text)
        self.assertIn('data-copy-value="2050292584830"', stock_page.text)

        copy_sources = {
            "static/decision-center.js": "copyIdentifier",
            "static/rnp-dashboard.js": "copyIdentifier",
            "static/stock-randomizer.js": "copyIdentifier",
            "static/unit-economics-1c.js": "data-copy-kind",
            "static/unit-economics-1c-report.js": "copyIdentifier",
            "templates/store_content.html": "CheckStockIdentifierCopy.html",
        }
        for relative_path, expected in copy_sources.items():
            with self.subTest(relative_path=relative_path):
                source = (root / relative_path).read_text(encoding="utf-8")
                self.assertIn(expected, source)

        escaped = copy_identifier('123"456', "Баркод")
        self.assertIn('data-copy-kind="Баркод"', escaped)
        self.assertIn('data-copy-value="123&quot;456"', escaped)
        self.assertIn('aria-label="Скопировать баркод 123&quot;456"', escaped)

    def test_cost_report_tables_have_filters_and_separate_store_marketplace_columns(self) -> None:
        db.record_operation(
            "rimili",
            "delivery",
            "manual",
            [
                {
                    "article": "949558341",
                    "barcode": "2050292584830",
                    "name": "Ретро гирлянда",
                    "quantity": 2,
                }
            ],
            1,
            "Unit Admin",
            NOW,
            to_fulfillment="ФулСервис Подольск",
            to_marketplace="WB",
        )
        for view, table_class in (
            ("summary", "cost-summary-table"),
            ("deliveries", "cost-operation-table"),
            ("fbs_sales", "cost-sales-table"),
        ):
            with self.subTest(view=view):
                response = self.client.get(
                    "/stock/cost-report",
                    params={
                        "date_from": "2026-08-11",
                        "date_to": "2026-08-13",
                        "view": view,
                    },
                )

                self.assertEqual(response.status_code, 200, response.text[:500])
                if view != "fbs_sales":
                    self.assertIn(f'class="{table_class} data-table"', response.text)
                    self.assertIn("<th>Магазин</th><th>Маркетплейс</th>", response.text)
                if view == "summary":
                    self.assertIn(
                        "<th>Продано FBS по формуле</th><th>ЗЦ продаж FBS</th>",
                        response.text,
                    )
                self.assertIn("Все маркетплейсы", response.text)
                self.assertNotIn("Все каналы", response.text)
                self.assertNotIn("нет ЗЦ для", response.text)

    def test_cost_report_can_classify_shipment_as_fbs_transfer(self) -> None:
        operation_id = db.record_operation(
            "rimili",
            "shipment",
            "manual",
            [
                {
                    "article": "949558341",
                    "barcode": "2050292584830",
                    "name": "Ретро гирлянда",
                    "quantity": 2,
                }
            ],
            1,
            "Unit Admin",
            NOW,
            from_fulfillment="ФулСервис Подольск",
            from_marketplace="WB",
        )

        response = self.client.post(
            f"/stock/cost-report/operations/{operation_id}/fbs-transfer",
            data={
                "is_fbs_transfer": "1",
                "date_from": "2026-08-11",
                "date_to": "2026-08-13",
                "view": "shipments",
            },
            follow_redirects=False,
        )

        self.assertEqual(response.status_code, 303, response.text)
        operations = db.get_operations_with_items_for_period(
            ("rimili",),
            "2026-08-11T00:00:00+00:00",
            "2026-08-13T23:59:59+00:00",
        )
        shipment = next(operation for operation in operations if operation["id"] == operation_id)
        self.assertEqual(shipment["is_fbs_transfer"], 1)

    def test_unit_economics_1c_cabinet_settings_page_and_api(self) -> None:
        root = Path(__file__).resolve().parents[2]
        settings_script = (root / "static" / "unit-economics-1c-cabinet-settings.js").read_text(
            encoding="utf-8"
        )
        settings_styles = (root / "static" / "unit-economics-1c-cabinet-settings.css").read_text(
            encoding="utf-8"
        )
        redirect = self.client.get("/sales/unit-economics-1c/cabinet-settings", follow_redirects=False)
        self.assertEqual(redirect.status_code, 303)
        self.assertEqual(redirect.headers["location"], "/admin/integrations#cabinet-settings")
        page = self.client.get("/admin/integrations")
        self.assertLess(page.text.index("API-ключи маркетплейсов"), page.text.index('id="cabinet-settings"'))
        self.assertLess(page.text.index('id="cabinet-settings"'), page.text.index("<h2>Все выгрузки</h2>"))
        self.assertEqual(page.status_code, 200)
        self.assertIn("Ввод данных по кабинетам", page.text)
        self.assertIn("Комиссия команды обновляется ночью", page.text)
        self.assertNotIn("Только хранение данных", page.text)
        self.assertIn("unit-economics-1c-cabinet-settings.js", page.text)
        self.assertIn('id="ue1cs-source-sync"', page.text)
        self.assertIn("Выгрузить себес", page.text)
        self.assertIn('id="ue1cs-price-sync"', page.text)
        self.assertIn("Выгрузить цены", page.text)
        self.assertIn('"acquiring_percent": 3.8', page.text)
        self.assertIn('"buyout_period_days": 14', page.text)
        self.assertIn('"team_commission_percent": 0.0', page.text)
        self.assertIn('"vat_percent": 9.0', page.text)
        self.assertIn('"usn_percent": 0.0', page.text)
        self.assertIn('"tax_system": "usn"', page.text)
        self.assertIn("Google Sheets", page.text)
        self.assertIn("key: 'vat_percent', label: 'Налог НДС'", settings_script)
        self.assertIn("key: 'buyout_period_days', label: 'Период расчёта'", settings_script)
        self.assertIn("Введите целое число от 1 до 29.", settings_script)
        self.assertIn("key: 'usn_percent', label: 'Налог УСН'", settings_script)
        self.assertIn("key: 'osno_percent', label: 'Налог ОСНО'", settings_script)
        self.assertIn("key: 'tax_system', label: 'Система налогообложения'", settings_script)
        self.assertIn("gogolOnly: true", settings_script)
        self.assertIn("payload.usn_percent = 0", settings_script)
        self.assertIn("payload.osno_percent = 0", settings_script)
        self.assertIn(".ue1cs-field[hidden]", settings_styles)

        payload = {
            "buyout_period_days": 21,
            "default_buyout_percent": 85.5,
            "acceptance_coefficient": 1.2,
            "wb_extra_tariff_percent": 2.3,
            "acquiring_percent": 4.2,
            "vat_percent": 10,
            "usn_percent": 6,
            "osno_percent": 0,
            "tax_system": "usn",
        }
        saved = self.client.put(
            "/api/unit-economics-1c/cabinet-settings/rimili",
            json=payload,
            headers={"X-Requested-With": "fetch"},
        )
        self.assertEqual(saved.status_code, 200, saved.text)
        self.assertEqual(saved.json()["settings"]["buyout_period_days"], 21)
        self.assertEqual(saved.json()["settings"]["acquiring_percent"], 4.2)
        self.assertEqual(saved.json()["settings"]["team_commission_percent"], 0)
        self.assertEqual(saved.json()["settings"]["vat_percent"], 10)
        self.assertEqual(saved.json()["settings"]["usn_percent"], 6)
        self.assertEqual(saved.json()["settings"]["tax_system"], "usn")

        loaded = self.client.get("/api/unit-economics-1c/cabinet-settings?marketplace=WB")
        self.assertEqual(loaded.status_code, 200)
        rimili = next(item for item in loaded.json()["items"] if item["store_slug"] == "rimili")
        self.assertEqual(rimili["acquiring_percent"], 4.2)
        self.assertEqual(rimili["buyout_period_days"], 21)
        self.assertEqual(rimili["default_buyout_percent"], 85.5)
        self.assertEqual(rimili["team_commission_percent"], 0)
        self.assertEqual(rimili["vat_percent"], 10)
        self.assertEqual(rimili["usn_percent"], 6)
        gogol_payload = {
            **payload,
            "vat_percent": 20,
            "osno_percent": 25,
            "tax_system": "osno",
        }
        gogol_saved = self.client.put(
            "/api/unit-economics-1c/cabinet-settings/gogol",
            json=gogol_payload,
            headers={"X-Requested-With": "fetch"},
        )
        self.assertEqual(gogol_saved.status_code, 200, gogol_saved.text)
        self.assertEqual(gogol_saved.json()["settings"]["vat_percent"], 20)
        self.assertEqual(gogol_saved.json()["settings"]["usn_percent"], 0)
        self.assertEqual(gogol_saved.json()["settings"]["osno_percent"], 25)
        self.assertEqual(gogol_saved.json()["settings"]["tax_system"], "osno")
        gogol_usn_saved = self.client.put(
            "/api/unit-economics-1c/cabinet-settings/gogol",
            json={**gogol_payload, "usn_percent": 5, "tax_system": "usn"},
            headers={"X-Requested-With": "fetch"},
        )
        self.assertEqual(gogol_usn_saved.status_code, 200, gogol_usn_saved.text)
        self.assertEqual(gogol_usn_saved.json()["settings"]["usn_percent"], 5)
        self.assertEqual(gogol_usn_saved.json()["settings"]["osno_percent"], 0)
        gogol_loaded = self.client.get("/api/unit-economics-1c/cabinet-settings?marketplace=WB")
        gogol = next(item for item in gogol_loaded.json()["items"] if item["store_slug"] == "gogol")
        self.assertEqual(gogol["usn_percent"], 5)
        self.assertEqual(gogol["osno_percent"], 0)
        for invalid_percent in (0, -1, 100.01):
            invalid = self.client.put(
                "/api/unit-economics-1c/cabinet-settings/rimili",
                json={**payload, "default_buyout_percent": invalid_percent},
                headers={"X-Requested-With": "fetch"},
            )
            self.assertEqual(invalid.status_code, 422)
        for invalid_days in (0, 30, 1.5):
            invalid = self.client.put(
                "/api/unit-economics-1c/cabinet-settings/rimili",
                json={**payload, "buyout_period_days": invalid_days},
                headers={"X-Requested-With": "fetch"},
            )
            self.assertEqual(invalid.status_code, 422)
        unit_product = self._unit_economics_product("rimili", "949558341")
        self.assertEqual(unit_product["details"]["buyout_percent"], 85.5)
        self.assertEqual(unit_product["details"]["acquiring"], 4.2)
        self.assertEqual(unit_product["details"]["team_commission_percent"], 0)
        self.assertEqual(unit_product["details"]["vat_percent"], 10)
        self.assertEqual(unit_product["details"]["usn_percent"], 6)
        forbidden = self.client.put(
            "/api/unit-economics-1c/cabinet-settings/rimili",
            json={**payload, "team_commission_percent": 2.8},
            headers={"X-Requested-With": "fetch"},
        )
        self.assertEqual(forbidden.status_code, 422)
        self.assertEqual(
            self.client.put(
                "/api/unit-economics-1c/cabinet-settings/missing",
                json=payload,
                headers={"X-Requested-With": "fetch"},
            ).status_code,
            404,
        )
        invalid = {**payload, "vat_percent": -1}
        self.assertEqual(
            self.client.put(
                "/api/unit-economics-1c/cabinet-settings/rimili",
                json=invalid,
                headers={"X-Requested-With": "fetch"},
            ).status_code,
            422,
        )

    def test_unit_economics_1c_manual_source_sync(self) -> None:
        report = {
            "ok": True,
            "saved": 945,
            "sheet_count": 6,
            "source_rows": 1791,
            "matched": 945,
            "unmatched": 846,
        }
        with mock.patch.object(
            unit_economics.unit_economics_1c_source,
            "sync_all",
            return_value=report,
        ) as sync:
            response = self.client.post(
                "/api/unit-economics-1c/source-data/sync",
                headers={"X-Requested-With": "fetch"},
            )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertTrue(response.json()["ok"])
        self.assertEqual(response.json()["report"]["saved"], 945)
        self.assertTrue(response.json()["items"])
        sync.assert_called_once_with()

        with mock.patch.object(
            unit_economics.unit_economics_1c_source,
            "sync_all",
            side_effect=unit_economics.unit_economics_1c_source.SourceDataError("нет доступа"),
        ):
            failed = self.client.post(
                "/api/unit-economics-1c/source-data/sync",
                headers={"X-Requested-With": "fetch"},
            )
        self.assertEqual(failed.status_code, 502)
        self.assertIn("нет доступа", failed.json()["error"])

    def test_unit_economics_1c_uses_sidebar_marketplace_navigation(self) -> None:
        response = self.client.get("/sales/unit-economics-1c")
        products = self._unit_economics_products()

        self.assertEqual(response.status_code, 200)
        self.assertIn("Юнит-экономика 1С", response.text)
        self.assertNotIn("949558341", response.text)
        self.assertTrue(any(item["article"] == "949558341" for item in products))
        summary = next(item for item in products if item["article"] == "949558341")
        self.assertTrue(summary["name"])
        self.assertEqual(set(summary["price"]), {"current", "with_spp"})
        self.assertNotIn("details", summary)
        self.assertNotIn("history", summary)
        detail = self._unit_economics_product("rimili", "949558341")
        self.assertIn("details", detail)
        self.assertEqual(len(detail["history"]), 21)
        self.assertIn('id="ue1c-product-rows"', response.text)
        self.assertIn('id="ue1c-products-loading"', response.text)
        self.assertIn('"products": []', response.text)
        self.assertIn('id="ue1c-store-filter"', response.text)
        self.assertIn('id="ue1c-detail"', response.text)
        self.assertIn('id="ue1c-chart"', response.text)
        self.assertIn('id="ue1c-pagination"', response.text)
        self.assertIn('id="ue1c-page-size"', response.text)
        self.assertIn('<option value="100">100</option>', response.text)
        self.assertIn('id="ue1c-table-head"', response.text)
        self.assertIn('id="ue1c-columns-toggle"', response.text)
        self.assertIn('id="ue1c-column-list"', response.text)
        self.assertNotIn('data-state-filter="risk"', response.text)
        self.assertNotIn('data-state-filter="low"', response.text)
        self.assertIn("Товары в минус", response.text)
        self.assertIn("Новинки", response.text)
        self.assertNotIn('class="ue1c-col-price"', response.text)
        self.assertIn('class="data-table ue1c-redesign-table"', response.text)
        self.assertIn('id="ue1c-column-list"', response.text)
        self.assertNotIn("Товарная юнит-экономика", response.text)
        self.assertNotIn('class="ue1c-kpis"', response.text)
        self.assertIn('href="/sales/unit-economics-1c/ozon"', response.text)
        self.assertNotIn('class="ue1c-market-tabs"', response.text)

        ozon = self.client.get("/sales/unit-economics-1c/ozon")
        yandex = self.client.get("/sales/unit-economics-1c/yandex-market")
        self.assertEqual(ozon.status_code, 200)
        self.assertEqual(yandex.status_code, 200)
        self.assertIn("Раздел Ozon в разработке", ozon.text)
        self.assertIn("Раздел Яндекс Маркета в разработке", yandex.text)
        self.assertIn('class="nav-subitem active" href="/sales/unit-economics-1c/ozon"', ozon.text)
        self.assertIn(
            'class="nav-subitem active" href="/sales/unit-economics-1c/yandex-market"',
            yandex.text,
        )

    def test_unit_economics_1c_accepts_closed_period_length(self) -> None:
        with mock.patch.object(
            unit_economics.unit_economics_1c,
            "load_product_metrics",
            return_value={},
        ) as load_metrics:
            response = self.client.get(
                "/sales/unit-economics-1c",
                params={"data": "1", "period_days": "14"},
            )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["period_days"], 14)
        self.assertEqual(
            (date.fromisoformat(payload["period_to"]) - date.fromisoformat(payload["period_from"])).days,
            13,
        )
        self.assertEqual(load_metrics.call_args_list[0].kwargs["period_days"], 14)

    def test_unit_economics_1c_accepts_custom_closed_date_range(self) -> None:
        with mock.patch.object(
            unit_economics.unit_economics_1c,
            "load_product_metrics",
            return_value={},
        ) as load_metrics:
            response = self.client.get(
                "/sales/unit-economics-1c",
                params={
                    "data": "1",
                    "date_from": "2026-08-10",
                    "date_to": "2026-08-19",
                },
            )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["period_mode"], "custom")
        self.assertEqual(payload["period_days"], 10)
        self.assertEqual(payload["period_from"], "2026-08-10")
        self.assertEqual(payload["period_to"], "2026-08-19")
        self.assertEqual(load_metrics.call_args_list[0].kwargs["period_days"], 10)
        self.assertEqual(load_metrics.call_args_list[0].kwargs["today"], date(2026, 8, 19))

    def test_unit_economics_1c_rejects_invalid_custom_date_range(self) -> None:
        response = self.client.get(
            "/sales/unit-economics-1c",
            params={
                "data": "1",
                "date_from": "2026-08-20",
                "date_to": "2026-08-19",
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("начала периода", response.json()["error"])

    def test_unit_economics_1c_ignores_legacy_catalog_exclusions(self) -> None:
        connection = core.get_connection()
        connection.execute(
            """
            CREATE TABLE catalog_product_exclusions (
                store_slug TEXT NOT NULL,
                marketplace TEXT NOT NULL,
                nm_id TEXT NOT NULL,
                status TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (store_slug, marketplace, nm_id)
            )
            """
        )
        connection.execute(
            """
            INSERT INTO catalog_product_exclusions
                (store_slug, marketplace, nm_id, status, updated_at)
            VALUES ('rimili', 'WB', '949558341', 'Старье', ?)
            """,
            (NOW,),
        )
        connection.commit()
        connection.close()

        products = self._unit_economics_products()

        self.assertTrue(any(item["article"] == "949558341" for item in products))

    def test_unit_economics_1c_product_contains_only_real_or_null_values(self) -> None:
        product = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {
                "article": "949558341",
                "barcode": "2050292584830",
                "name": "Ретро гирлянда",
                "image_url": "",
                "fbs_stock": 12,
                "fbo_stock": 8,
                "ff_available": 5,
            },
            {
                "seller_base_price": 10000,
                "retail_price": 7950,
                "club_discounted_price": 7600,
                "customer_price_with_spp": 5756,
                "customer_price_with_wallet": 5640,
                "day": "2026-08-19",
                "customer_price_window_days": 7,
                "customer_price_orders_count": 3,
            },
        )

        self.assertEqual(len(product["history"]), 7)
        self.assertEqual(product["price"]["current"], 7950)
        self.assertEqual(product["price"]["base"], 10000)
        self.assertEqual(product["price"]["club"], 7600)
        self.assertEqual(product["price"]["with_spp"], 5756)
        self.assertEqual(product["price"]["with_wallet"], 5640)
        self.assertEqual(product["price"]["window_days"], 7)
        self.assertIn("margin_rub", product["history"][0])
        self.assertNotIn("margin_percent", product["history"][0])
        self.assertIn("advertising_rub", product["history"][0])
        self.assertIn("drr_percent", product["history"][0])
        self.assertIn("purchase_value", product["history"][0])
        self.assertIn("purchase_cost", product["details"])
        self.assertEqual(product["details"]["acquiring"], 3.8)
        self.assertNotIn("spp_percent", product["details"])
        self.assertIsNone(product["rating"])
        self.assertTrue(all(value is None for value in product["tag_data"].values()))
        self.assertEqual(product["stock"]["fbs"], 12)
        self.assertEqual(product["stock"]["fbo"], 8)
        self.assertEqual(product["stock"]["fulfillment"], 5)
        self.assertEqual(product["stock"]["total"], 25)
        self.assertEqual(product["stock"]["days"], 0)
        self.assertEqual(product["stock"]["orders_21d"], 0)
        self.assertEqual(product["stock"]["average_daily_orders"], 0)
        self.assertEqual(product["history"][0]["margin_rub"], 0)
        self.assertEqual(product["history"][0]["drr_percent"], 0)
        self.assertIsNone(product["history"][0]["fbs_units"])
        self.assertEqual(product["history"][-1]["fbs_units"], 12)
        self.assertIsNone(product["details"]["subject"])
        self.assertEqual(product["details"]["commission_percent"], 0)
        self.assertEqual(product["details"]["purchase_cost"], 0)
        self.assertEqual(product["details"]["fulfillment_cost"], 0)
        self.assertNotIn("fbo_commission", product["details"])
        self.assertNotIn("fbs_commission", product["details"])
        self.assertNotIn("tax_type", product["details"])
        self.assertEqual(product["details"]["vat_percent"], 0)
        self.assertEqual(product["details"]["usn_percent"], 0)
        self.assertEqual(product["details"]["osno_percent"], 0)
        self.assertEqual(product["details"]["tax_system"], "usn")
        self.assertNotIn("margin_percent", product["details"])
        self.assertNotIn("fixed_margin", product["details"])
        self.assertNotIn("package", product["details"])

        product_not_for_sale = unit_economics._unit_economics_1c_mock_product(
            "tris",
            {"article": "1449522671", "name": "Новый товар"},
            product_metrics={
                "period_from": "2026-08-20",
                "period_to": "2026-08-26",
                "period_days": 7,
                "orders_amount": 0,
                "orders_count": 0,
                "spend": 0,
                "daily": [],
            },
            product_reference={"purchase_price": 22_008.75},
        )
        self.assertIsNone(product_not_for_sale["price"]["current"])
        self.assertIsNone(product_not_for_sale["price"]["with_spp"])
        self.assertIsNone(product_not_for_sale["current_economics"]["margin"])
        self.assertIsNone(product_not_for_sale["current_economics"]["roi"])
        self.assertIsNone(product_not_for_sale["economics_7d"]["margin"])
        self.assertIsNone(product_not_for_sale["economics_7d"]["roi"])

        product_with_demand = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {
                "article": "949558341",
                "fbs_stock": None,
                "fbo_stock": 8,
                "ff_available": None,
            },
            stock_order_metrics={
                "period_from": "2026-07-30",
                "period_to": "2026-08-19",
                "period_days": 21,
                "orders_count": 42,
            },
        )
        self.assertEqual(product_with_demand["stock"]["fbs"], 0)
        self.assertEqual(product_with_demand["stock"]["fbo"], 8)
        self.assertEqual(product_with_demand["stock"]["fulfillment"], 0)
        self.assertEqual(product_with_demand["stock"]["total"], 8)
        self.assertEqual(product_with_demand["stock"]["average_daily_orders"], 2)
        self.assertEqual(product_with_demand["stock"]["days"], 4)

        product_with_stock_history = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {"article": "949558341", "fbs_stock": 12, "fbo_stock": 8, "ff_available": 5},
            product_metrics={"period_to": "2026-08-19", "daily": []},
            stock_history_by_day={
                "2026-08-18": {"fbs": 9, "fbo": 7, "fulfillment": 4},
                "2026-08-19": {"fbs": 11, "fbo": 6, "fulfillment": 3},
            },
        )
        self.assertEqual(product_with_stock_history["history"][-2]["fbs_units"], 9)
        self.assertEqual(product_with_stock_history["history"][-1]["fbs_units"], 11)
        self.assertEqual(product_with_stock_history["history"][-1]["fbo_units"], 6)
        self.assertEqual(product_with_stock_history["history"][-1]["fulfillment_units"], 3)

        product_with_reputation_and_glue = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {"article": "949558341", "name": "Основной товар"},
            reputation={"rating": 4.8, "reviews_count": 123},
            glued_products=[{"article": "949558342", "name": "Связанный товар"}],
        )
        self.assertEqual(product_with_reputation_and_glue["rating"], 4.8)
        self.assertEqual(product_with_reputation_and_glue["reviews_count"], 123)
        self.assertEqual(
            product_with_reputation_and_glue["glued_products"],
            [{"article": "949558342", "name": "Связанный товар"}],
        )

        new_product = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {"article": "949558343", "name": "Новинка"},
            product_reference={"card_created_at": "2026-08-01T10:15:00Z"},
            first_sale_at="2026-08-10",
            sales_age_today=datetime(2026, 8, 19, tzinfo=UTC).date(),
        )
        self.assertTrue(new_product["is_new"])
        self.assertEqual(new_product["sales_days"], 10)
        self.assertEqual(new_product["sales_started_at"], "2026-08-10")

        old_product = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {"article": "949558344", "name": "Старый товар"},
            first_sale_at="2026-06-01",
            sales_age_today=datetime(2026, 8, 19, tzinfo=UTC).date(),
        )
        self.assertFalse(old_product["is_new"])
        self.assertEqual(old_product["sales_days"], 80)

    def test_unit_economics_1c_glued_products_are_interactive_and_located_in_table(self) -> None:
        root = Path(__file__).resolve().parents[2]
        script = (root / "static" / "unit-economics-1c.js").read_text(encoding="utf-8")
        styles = (root / "static" / "unit-economics-1c.css").read_text(encoding="utf-8")
        template = (root / "templates" / "unit_economics_1c_content.html").read_text(
            encoding="utf-8"
        )

        self.assertLess(template.index("Динамика за 14 дней"), template.index("Связанные артикулы"))
        self.assertNotIn("склейка WB", template)
        self.assertNotIn("заказы столбиками ×10, показатели линиями", template)
        self.assertNotIn("Цель: цена", script)
        self.assertIn('class="ue1c-calculator ue1c-drawer-section"', template)
        self.assertIn("overflow-x: hidden", styles)
        self.assertIn('data-copy-tooltip="нажмите чтобы скопировать"', script)
        self.assertIn('data-glued-product-open="', script)
        self.assertIn("function locateProductInTable(product)", script)
        self.assertIn("Math.floor(productIndex / state.pageSize) + 1", script)
        self.assertIn("targetRow.scrollIntoView", script)
        self.assertIn("targetRow.classList.add('is-located')", script)
        self.assertIn("}, 5000);", script)
        self.assertIn("tr.is-located td", styles)
        self.assertIn("function coverageTitle(label, coverage)", script)
        self.assertIn("function coverageValue(label, value, formatter, coverage, suffix)", script)
        self.assertIn("coverageCellClass(marginCoverage)", script)
        self.assertIn("function negativeValueClass(value)", script)
        self.assertIn("ue1c-roi-negative", script)
        self.assertIn("td.ue1c-roi-negative", styles)
        self.assertIn("var AUTO_REFRESH_INTERVAL_MS = 5 * 60 * 1000", script)
        self.assertIn("function refreshProductsIfStale(force)", script)
        self.assertIn("loadProducts({ silent: true, refreshDetail: true })", script)
        self.assertIn("document.addEventListener('visibilitychange'", script)
        self.assertIn("async function refreshSelectedDetail()", script)
        self.assertIn('id="ue1c-period-from" type="date"', template)
        self.assertIn('id="ue1c-period-to" type="date"', template)
        self.assertIn('id="ue1c-period-apply"', template)
        self.assertIn('<option value="custom">Свои даты</option>', template)
        self.assertNotIn('id="ue1c-refresh"', template)
        self.assertNotIn("nodes.refresh", script)
        self.assertIn("query.set('date_from', state.periodFrom)", script)
        self.assertIn("query.set('date_to', state.periodTo)", script)
        self.assertIn("if (group.key === 'advertising' && customRange)", script)
        self.assertIn("ue1c-partial-cell", styles)
        self.assertIn(".ue1c-coverage-value.is-partial", styles)

    def test_unit_economics_1c_commission_panel_uses_split_taxes(self) -> None:
        root = Path(__file__).resolve().parents[2]
        script = (root / "static" / "unit-economics-1c.js").read_text(encoding="utf-8")
        template = (root / "templates" / "unit_economics_1c_content.html").read_text(encoding="utf-8")

        self.assertNotIn("'NULL'", script)
        self.assertNotIn("parameter('FBO'", script)
        self.assertNotIn("parameter('FBS'", script)
        self.assertNotIn("parameter('Ставка'", script)
        self.assertNotIn("parameter('Тип'", script)
        self.assertNotIn("parameter('ИРП'", script)
        self.assertNotIn("parameter('Упаковка'", script)
        self.assertNotIn("parameter('Фикс. маржа'", script)
        self.assertIn("parameter('Налоговая система'", script)
        self.assertIn("parameter('НДС', nullable(item.vat_percent, decimal, '%'))", script)
        self.assertIn("metric('Чистая прибыль', nullable(metrics.margin, preciseMoney)", script)
        self.assertNotIn("metric('Маржинальность'", script)
        for key in (
            "retail",
            "spp",
            "client",
            "walletPercent",
            "wallet",
            "commission",
            "commissionRub",
            "drr",
            "advertisingRub",
            "logistics",
            "storage",
            "storageTotal",
            "acquiringPercent",
            "acquiringRub",
            "purchase",
            "team",
            "teamRub",
            "fulfillment",
            "vat",
            "vatRub",
            "usn",
            "osno",
            "secondaryTaxRub",
        ):
            self.assertIn(f'data-calculator-input="{key}"', template)
        self.assertIn("current * acquiringPercent / 100", script)
        self.assertNotIn("walletPrice * acquiringPercent / 100", script)
        self.assertIn("- acquiringPercent / 100", script)
        self.assertNotIn("- walletPriceFactor * acquiringPercent / 100", script)
        self.assertNotIn("current * drrPercent / 100", script)
        self.assertIn("var advertising = advertisingRub", script)
        self.assertIn("advertisingBase * drr / 100", script)
        self.assertIn("advertisingRub / advertisingBase * 100", script)
        self.assertNotIn("buyoutRatio", script)
        self.assertIn("current * teamCommissionPercent / 100", script)
        self.assertNotIn("clientPrice * teamCommissionPercent / 100", script)
        self.assertIn("- teamCommissionPercent / 100", script)
        self.assertIn("clientPrice * vatPercent / (100 + vatPercent)", script)
        self.assertIn("(clientPrice - vat) * activeTaxPercent / 100", script)
        self.assertIn("clientPrice * activeTaxPercent / 100", script)
        self.assertIn("clientPriceFactor * (vatFactor + secondaryTaxFactor)", script)
        self.assertNotIn("taxRate", script)
        self.assertIn("storageRate * turnoverDays", script)
        self.assertIn("current - acquiring - logistics - storage - commission - advertising", script)
        self.assertIn("netRevenue - purchase - fulfillment - teamCommission - tax", script)
        self.assertIn("margin / purchase * 100", script)
        self.assertIn("function sppPriceFactor(product)", script)
        self.assertIn("function walletPriceFactor(product)", script)
        self.assertIn("function syncLinkedPriceInputs(product, source)", script)
        self.assertIn("source === 'retail'", script)
        self.assertIn("source === 'client'", script)
        self.assertIn("source === 'wallet'", script)
        self.assertIn("syncLinkedPriceInputs(product, editedPriceKind)", script)
        self.assertIn('id="ue1c-spp-price-input" data-calculator-input="client"', template)
        self.assertIn('id="ue1c-calculator-mode" type="checkbox" role="switch"', template)
        self.assertIn('id="ue1c-calculator-inputs"', template)
        self.assertIn("Затраты на ФФ", template)
        self.assertIn("Рекламные расходы на выкупленную единицу, руб", template)
        self.assertIn('id="ue1c-secondary-tax-label"', template)
        for output_id in (
            "ue1c-spp-percent",
            "ue1c-wallet-percent",
            "ue1c-commission-rub",
            "ue1c-advertising-rub",
            "ue1c-acquiring-percent",
            "ue1c-acquiring-rub",
            "ue1c-storage-rub",
            "ue1c-purchase-rub",
            "ue1c-team-rub",
            "ue1c-vat-rub",
            "ue1c-secondary-tax-rub",
        ):
            self.assertIn(f'id="{output_id}"', template)
        self.assertIn("calculatorFields.classList.toggle('is-expanded'", script)
        self.assertIn("function syncDetailedCalculatorInputs(product, source)", script)
        self.assertIn("syncPair('commission', 'commissionRub'", script)
        self.assertIn("syncPair('acquiringPercent', 'acquiringRub'", script)
        self.assertIn("syncPair('team', 'teamRub'", script)
        self.assertIn("finite(values.storageTotal", script)
        self.assertIn("finite(values.vatRub", script)
        self.assertIn("finite(values.secondaryTaxRub", script)
        self.assertIn("product.store_slug === 'gogol'", script)
        self.assertNotIn("data-price-filter=", template)
        self.assertNotIn("data-color-filter", template)
        self.assertNotIn("is-price-queued", script)
        self.assertNotIn('id="ue1c-reset-prices"', template)
        self.assertNotIn('id="ue1c-show-changed"', template)
        self.assertNotIn("priceQueueTtlMs", script)
        self.assertNotIn("schedulePriceQueueExpiry", script)
        self.assertNotIn("expirePendingPrices", script)
        self.assertNotIn("Выгрузить цены", template)
        self.assertIn('id="ue1c-price-confirm-modal"', template)
        self.assertIn("/api/unit-economics-1c/prices/preview", script)
        self.assertIn("Подтвердить и отправить", template)
        self.assertEqual(
            unit_economics.unit_economics_1c_prices.PRICE_SYNC_AFTER_UPLOAD_DELAY_SECONDS,
            10.0,
        )
        self.assertNotIn("ue1c-store-row", script)
        self.assertIn("function calculateSppPercent(product)", script)
        self.assertIn("(withoutSpp - withSpp) / withoutSpp * 100", script)
        self.assertIn("{ index: 24, label: 'СПП, %'", script)
        self.assertIn("nullable(currentSpp, decimal, '%')", script)
        self.assertIn(
            "parameter('СПП', nullable(calculateSppPercent(product), decimal, '%'))",
            script,
        )
        self.assertNotIn("item.spp_percent", script)
        self.assertNotIn("Структура расходов", template)
        self.assertNotIn('id="ue1c-costs"', template)
        self.assertNotIn("nodes.costs", script)
        self.assertIn('<details class="ue1c-formula-popover">', template)
        self.assertIn("Показать формулу чистой прибыли", template)
        self.assertIn(
            "чистая выручка − закупочная цена − фулфилмент − комиссия компании − НДС − УСН/ОСНО",
            template,
        )
        self.assertIn(
            "цена без СПП − эквайринг − логистика − хранение − комиссия WB − реклама на единицу",
            template,
        )
        self.assertIn(
            "Реклама на единицу</strong> = рекламные расходы за 7 дней ÷ заказы воронки за 7 дней ÷ (процент выкупа ÷ 100)",
            template,
        )
        self.assertIn(
            "ДРР с выкупом</strong> = рекламные расходы за 7 дней ÷ (оборот заказов за 7 дней × процент выкупа ÷ 100) × 100%",
            template,
        )
        self.assertIn("Рекламные расходы на выкупленную единицу, руб", template)
        self.assertIn("НДС</strong> = цена с СПП ÷ (100 + ставка) × ставка", template)
        self.assertIn("УСН</strong> = (цена покупателя − НДС) × ставка ÷ 100", template)
        self.assertIn('id="ue1c-subject-select"', template)
        self.assertIn('type="search" list="ue1c-subject-options"', template)
        self.assertIn('id="ue1c-subject-options"', template)
        self.assertNotIn("Симулятор параметров", template)
        self.assertNotIn("Параметр</span><span>Значение", template)

    def test_unit_economics_1c_history_calculates_daily_drr_from_raw_metrics(self) -> None:
        product = unit_economics._unit_economics_1c_mock_product(
            "trusthome",
            {"article": "551394618", "name": "Товар"},
            product_metrics={
                "period_to": "2026-08-19",
                "period_days": 7,
                "buyout_percent": 80,
                "daily": [
                    {
                        "date": "2026-08-15",
                        "advertising_spend": 211.5,
                        "orders_amount": 4372,
                        "orders_count": 3,
                    }
                ],
            },
        )

        day = next(item for item in product["history"] if item["date"] == "2026-08-15")
        self.assertEqual(day["drr_percent"], 6.05)

    def test_unit_economics_1c_history_uses_report_daily_economics(self) -> None:
        product = unit_economics._unit_economics_1c_mock_product(
            "trusthome",
            {"article": "551394618", "name": "Товар"},
            product_metrics={
                "period_to": "2026-08-15",
                "period_days": 1,
                "buyout_percent": 80,
            },
            history_product_metrics={
                "period_to": "2026-08-15",
                "period_days": 1,
                "buyout_percent": 80,
                "daily": [
                    {
                        "date": "2026-08-15",
                        "advertising_spend": 211.5,
                        "orders_amount": 4372,
                        "orders_count": 3,
                    }
                ],
            },
            history_day_economics={
                "2026-08-15": {
                    "margin": 321.25,
                    "purchase_value": 150,
                    "buyout_percent": 50,
                }
            },
        )

        day = product["history"][0]
        self.assertEqual(day["orders_count"], 3)
        self.assertEqual(day["advertising_rub"], 211.5)
        self.assertEqual(day["buyout_percent"], 50)
        self.assertEqual(day["purchased_units"], 1.5)
        self.assertEqual(day["drr_percent"], 9.68)
        self.assertEqual(day["margin_rub"], 321.25)

    def test_unit_economics_1c_period_margin_is_unit_margin_times_funnel_orders(self) -> None:
        product = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {"article": "949735537", "name": "Товар"},
            {
                "retail_price": 1000,
                "customer_price_with_spp": 800,
                "customer_price_with_wallet": 700,
            },
            acquiring_percent=3.8,
            product_metrics={
                "period_from": "2026-08-13",
                "period_to": "2026-08-19",
                "period_days": 7,
                "orders_amount": 8000,
                "orders_count": 10,
                "sold_count": 8,
                "buyout_percent": 80,
                "spend": 800,
                "drr": 10,
                "daily": [
                    {
                        "date": "2026-08-19",
                        "advertising_spend": 800,
                        "orders_amount": 8000,
                        "orders_count": 10,
                        "sold_count": 8,
                        "drr": 10,
                    }
                ],
            },
            team_commission_percent=4,
            vat_percent=9,
            usn_percent=6,
            product_reference={
                "turnover_days": 21,
                "purchase_price": 300,
                "fulfillment_cost": 50,
                "subject_commission_percent": 10,
            },
        )

        self.assertEqual(product["advertising"]["buyout_percent"], 80)
        self.assertEqual(product["advertising"]["average_daily_spend"], 114.29)
        self.assertEqual(product["advertising"]["spend_per_order"], 100)
        self.assertEqual(product["advertising"]["drr"], 12.5)
        self.assertEqual(product["economics_7d"]["turnover"], 8000)
        self.assertEqual(product["economics_7d"]["orders"], 10)
        self.assertEqual(product["details"]["vat_value"], 66.06)
        self.assertEqual(product["details"]["usn_value"], 44.04)
        self.assertEqual(product["details"]["tax_value"], 110.09)
        self.assertEqual(product["current_economics"]["margin"], 261.91)
        self.assertEqual(product["economics_7d"]["margin"], 2619.1)
        self.assertEqual(
            product["economics_7d"]["margin"],
            round(
                product["current_economics"]["margin"] * product["economics_7d"]["orders"],
                2,
            ),
        )
        self.assertEqual(product["economics_7d"]["roi"], 87.3)
        self.assertEqual(product["history"][-1]["purchased_units"], 8)
        self.assertEqual(product["history"][-1]["margin_rub"], 2095.28)

    def test_unit_economics_1c_separates_today_from_closed_seven_day_period(self) -> None:
        product = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {"article": "949735537", "name": "Товар"},
            {
                "retail_price": 1000,
                "customer_price_with_spp": 800,
            },
            product_metrics={
                "period_from": "2026-08-23",
                "period_to": "2026-08-29",
                "period_days": 7,
                "orders_amount": 7000,
                "orders_count": 7,
                "cancel_amount": 1000,
                "buyout_percent": 80,
                "spend": 700,
                "impressions": 1000,
                "clicks": 100,
                "ctr": 10,
                "cpc": 7,
                "daily": [],
            },
            current_product_metrics={
                "period_from": "2026-08-30",
                "period_to": "2026-08-30",
                "period_days": 1,
                "orders_amount": 1600,
                "orders_count": 2,
                "range_buyout_percent": 0,
                "buyout_percent": 75,
                "spend": 90,
                "daily": [],
            },
            closed_period_economics={
                "margin": 1234.56,
                "purchase_value": 1000,
                "roi": 123.46,
                "buyout_percent": 70,
                "complete": False,
                "coverage": {
                    "dates": ["2026-08-25", "2026-08-26", "2026-08-27", "2026-08-28", "2026-08-29"],
                    "days": 5,
                    "expected_days": 7,
                    "complete": False,
                    "period_from": "2026-08-25",
                    "period_to": "2026-08-29",
                    "missing_dates": ["2026-08-23", "2026-08-24"],
                },
            },
            turnover_coverage={
                "dates": [
                    "2026-08-24",
                    "2026-08-25",
                    "2026-08-26",
                    "2026-08-27",
                    "2026-08-28",
                    "2026-08-29",
                ],
                "days": 6,
                "expected_days": 7,
                "complete": False,
                "period_from": "2026-08-24",
                "period_to": "2026-08-29",
                "missing_dates": ["2026-08-23"],
            },
            product_reference={
                "purchase_price": 300,
                "fulfillment_cost": 50,
                "subject_commission_percent": 10,
            },
        )

        self.assertEqual(
            (product["economics_7d"]["period_from"], product["economics_7d"]["period_to"]),
            ("2026-08-23", "2026-08-29"),
        )
        self.assertEqual(product["economics_7d"]["turnover"], 6000)
        self.assertEqual(product["economics_7d"]["margin"], 1234.56)
        self.assertEqual(product["economics_7d"]["roi"], 123.46)
        self.assertEqual(product["economics_7d"]["turnover_coverage"]["days"], 6)
        self.assertEqual(product["economics_7d"]["margin_coverage"]["days"], 5)
        self.assertEqual(product["economics_7d"]["roi_coverage"]["days"], 5)
        self.assertEqual(product["advertising"]["period_from"], "2026-08-23")
        self.assertEqual(product["advertising"]["period_to"], "2026-08-29")
        self.assertEqual(product["advertising"]["spend"], 700)
        self.assertEqual(product["advertising"]["drr"], 12.5)
        self.assertEqual(product["advertising"]["spend_per_order"], 125)
        self.assertEqual(
            (product["current_economics"]["period_from"], product["current_economics"]["period_to"]),
            ("2026-08-30", "2026-08-30"),
        )
        self.assertEqual(product["current_economics"]["margin"], 455.75)
        self.assertEqual(product["current_economics"]["roi"], 151.92)
        self.assertEqual(product["current_economics"]["orders"], 2)
        self.assertEqual(product["current_economics"]["advertising_spend"], 90)
        self.assertEqual(product["details"]["buyout_percent"], 80)
        self.assertEqual(product["details"]["advertising_per_unit"], 56.25)

    def test_unit_profit_report_sums_margin_and_orders_for_each_day(self) -> None:
        result = unit_economics._report_historical_economics(
            date_from=date(2026, 8, 23),
            date_to=date(2026, 8, 25),
            daily_orders={
                "2026-08-23": {"orders_count": 2, "buyout_percent": 50},
                "2026-08-24": {"orders_count": 3, "buyout_percent": 80},
                "2026-08-25": {"orders_count": 4, "buyout_percent": 100},
            },
            margin_snapshots={
                "2026-08-23": {"unit_margin": 100, "purchase_price": 50, "inputs_json": json.dumps({"buyout_percent": 50})},
                "2026-08-24": {"unit_margin": 200, "purchase_price": 60, "inputs_json": json.dumps({"buyout_percent": 80})},
                "2026-08-25": {"unit_margin": 300, "purchase_price": 70, "inputs_json": json.dumps({"buyout_percent": 100})},
            },
            live_day=date(2026, 8, 28),
            live_unit_margin=None,
            live_purchase_price=None,
            daily_advertising={
                "2026-08-23": 10,
                "2026-08-24": 20,
                "2026-08-25": 30,
            },
        )

        self.assertEqual(result["margin"], 1720)
        self.assertEqual(result["purchase_value"], 474)
        self.assertEqual(result["orders"], 7.4)
        self.assertEqual(result["advertising_spend"], 60)
        self.assertTrue(result["complete"])

    def test_historical_economics_can_return_only_covered_days_for_dashboard(self) -> None:
        result = unit_economics._report_historical_economics(
            date_from=date(2026, 8, 23),
            date_to=date(2026, 8, 25),
            daily_orders={
                "2026-08-23": {"orders_count": 2, "buyout_percent": 50},
                "2026-08-24": {"orders_count": 3, "buyout_percent": 80},
                "2026-08-25": {"orders_count": 4, "buyout_percent": 100},
            },
            margin_snapshots={
                "2026-08-24": {"unit_margin": 200, "purchase_price": 60, "inputs_json": json.dumps({"buyout_percent": 80})},
                "2026-08-25": {"unit_margin": 300, "purchase_price": 70, "inputs_json": json.dumps({"buyout_percent": 100})},
            },
            live_day=date(2026, 8, 28),
            live_unit_margin=None,
            live_purchase_price=None,
            daily_advertising={
                "2026-08-23": 10,
                "2026-08-24": 20,
                "2026-08-25": 30,
            },
            allow_partial=True,
        )

        self.assertEqual(result["margin"], 1630)
        self.assertEqual(result["purchase_value"], 424)
        self.assertEqual(result["roi"], 384.43)
        self.assertEqual(result["orders"], 6.4)
        self.assertEqual(result["advertising_spend"], 50)
        self.assertFalse(result["complete"])
        self.assertEqual(result["missing_days"], ["2026-08-23"])
        self.assertEqual(result["coverage"]["dates"], ["2026-08-24", "2026-08-25"])
        self.assertEqual(result["coverage"]["days"], 2)
        self.assertEqual(result["coverage"]["expected_days"], 3)

        empty = unit_economics._report_historical_economics(
            date_from=date(2026, 8, 23),
            date_to=date(2026, 8, 23),
            daily_orders={"2026-08-23": {"orders_count": 1, "buyout_percent": 100}},
            margin_snapshots={},
            live_day=date(2026, 8, 28),
            live_unit_margin=None,
            live_purchase_price=None,
            allow_partial=True,
        )
        self.assertIsNone(empty["margin"])
        self.assertIsNone(empty["roi"])
        self.assertEqual(empty["coverage"]["days"], 0)

    def test_unit_profit_warning_lists_days_below_seventy_percent_coverage(self) -> None:
        rows = [
            {"store_slug": "rimili", "article": str(index)}
            for index in range(10)
        ]
        snapshots = {
            ("rimili", str(index)): {
                "2026-08-28": {},
                "2026-08-29": {},
            }
            for index in range(6)
        }
        for index in range(6, 7):
            snapshots[("rimili", str(index))] = {"2026-08-29": {}}

        missing_days = unit_economics._undercovered_margin_days(
            rows=rows,
            date_from=date(2026, 8, 28),
            date_to=date(2026, 8, 29),
            margin_snapshots_by_product=snapshots,
            daily_contexts={},
            live_day=date(2026, 9, 2),
        )

        self.assertEqual(missing_days, ["2026-08-28"])

    def test_unit_profit_warning_counts_live_snapshots_toward_coverage(self) -> None:
        rows = [
            {"store_slug": "rimili", "article": str(index)}
            for index in range(10)
        ]
        missing_days = unit_economics._undercovered_margin_days(
            rows=rows,
            date_from=date(2026, 9, 2),
            date_to=date(2026, 9, 2),
            margin_snapshots_by_product={},
            daily_contexts={
                ("rimili", str(index)): {"live_snapshot": {"unit_margin": 100}}
                for index in range(7)
            },
            live_day=date(2026, 9, 2),
        )

        self.assertEqual(missing_days, [])

    def test_unit_profit_report_never_substitutes_current_margin_for_missing_history(self) -> None:
        result = unit_economics._report_historical_economics(
            date_from=date(2026, 8, 23),
            date_to=date(2026, 8, 24),
            daily_orders={
                "2026-08-23": {"orders_count": 2, "buyout_percent": 100},
                "2026-08-24": {"orders_count": 3, "buyout_percent": 100},
            },
            margin_snapshots={
                "2026-08-24": {"unit_margin": 200, "purchase_price": 60},
            },
            live_day=date(2026, 8, 28),
            live_unit_margin=999,
            live_purchase_price=999,
        )

        self.assertIsNone(result["margin"])
        self.assertIsNone(result["purchase_value"])
        self.assertFalse(result["complete"])
        self.assertEqual(result["missing_days"], ["2026-08-23"])

    def test_unit_profit_report_allocates_daily_advertising_to_expected_buyouts(self) -> None:
        result = unit_economics._report_historical_economics(
            date_from=date(2026, 8, 28),
            date_to=date(2026, 8, 28),
            daily_orders={
                "2026-08-28": {"orders_count": 43, "buyout_percent": 80},
            },
            margin_snapshots={
                "2026-08-28": {
                    "unit_margin": 808.06,
                    "purchase_price": 1123.9,
                    "calculation_version": 2,
                    "inputs_json": json.dumps({"buyout_percent": 80}),
                },
            },
            live_day=date(2026, 8, 29),
            live_unit_margin=None,
            live_purchase_price=None,
            daily_advertising={"2026-08-28": 10_488},
        )

        self.assertEqual(result["orders"], 34.4)
        self.assertEqual(result["advertising_spend"], 10_488)
        self.assertEqual(result["margin"], 17_309.26)
        self.assertEqual(result["purchase_value"], 38_662.16)

    def test_unit_profit_report_uses_weekly_buyout_when_daily_value_is_zero(self) -> None:
        result = unit_economics._report_historical_economics(
            date_from=date(2026, 8, 28),
            date_to=date(2026, 8, 28),
            daily_orders={
                "2026-08-28": {"orders_count": 43, "buyout_percent": 0},
            },
            margin_snapshots={
                "2026-08-28": {
                    "unit_margin": 808.06,
                    "purchase_price": 1123.9,
                    "calculation_version": 2,
                    "inputs_json": json.dumps({"buyout_percent": 80}),
                },
            },
            live_day=date(2026, 8, 29),
            live_unit_margin=None,
            live_purchase_price=None,
            daily_advertising={"2026-08-28": 10_488},
            fallback_buyout_percent=78,
        )

        self.assertEqual(result["orders"], 34.4)
        self.assertEqual(result["margin"], 17_309.26)
        self.assertEqual(result["buyout_percent"], 80)

        gogol_taxes = unit_economics._unit_economics_1c_mock_product(
            "gogol",
            {"article": "GOGOL-TAX", "name": "Товар"},
            {"retail_price": 1000, "customer_price_with_spp": 800},
            vat_percent=20,
            usn_percent=6,
            osno_percent=25,
            tax_system="osno",
        )
        self.assertEqual(gogol_taxes["details"]["tax_system"], "osno")
        self.assertEqual(gogol_taxes["details"]["vat_value"], 133.33)
        self.assertEqual(gogol_taxes["details"]["usn_value"], 0)
        self.assertEqual(gogol_taxes["details"]["osno_value"], 200)
        self.assertEqual(gogol_taxes["details"]["tax_value"], 333.33)

        non_gogol_taxes = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {"article": "RIMILI-TAX", "name": "Товар"},
            {"retail_price": 1000, "customer_price_with_spp": 800},
            vat_percent=20,
            usn_percent=6,
            osno_percent=25,
            tax_system="osno",
        )
        self.assertEqual(non_gogol_taxes["details"]["tax_system"], "usn")
        self.assertEqual(non_gogol_taxes["details"]["vat_value"], 133.33)
        self.assertEqual(non_gogol_taxes["details"]["usn_value"], 40)
        self.assertEqual(non_gogol_taxes["details"]["osno_value"], 0)
        self.assertEqual(non_gogol_taxes["details"]["tax_value"], 173.33)

        storefront_only = unit_economics._unit_economics_1c_mock_product(
            "rimili",
            {"article": "949558341", "name": "Ретро гирлянда"},
            {"retail_price": None, "customer_price_with_spp": 2034},
        )
        self.assertIsNone(storefront_only["price"]["current"])
        self.assertEqual(storefront_only["price"]["with_spp"], 2034)
        self.assertNotIn("spp_percent", storefront_only["details"])

    def test_unit_profit_report_groups_categories_with_ratio_totals(self) -> None:
        rows = [
            {
                "subject": "Игрушки",
                "store_slug": "rimili",
                "store_name": "RIMILI",
                "manager": "Менеджер",
                "orders_count": 10,
                "orders_amount": 1000,
                "cancel_count": 2,
                "cancel_amount": 200,
                "net_orders_count": 8,
                "net_orders_amount": 800,
                "buyout_percent": 80,
                "buyout_orders_count": 10,
                "expected_buyout_amount": 800,
                "stock": 3,
                "stock_fbs": 1,
                "stock_fbo": 2,
                "stock_fulfillment": 0,
                "stock_days": 3,
                "stock_average_daily_orders": 1,
                "impressions": 100,
                "clicks": 10,
                "advertising_spend": 100,
                "margin_orders_count": 8,
                "margin": 400,
                "purchase_value": 800,
                "margin_complete": True,
                "margin_missing_days": [],
                "daily_calculations": [
                    {
                        "date": "2026-08-30",
                        "available": True,
                        "snapshot_available": True,
                        "orders_count": 10,
                        "net_orders_count": 8,
                        "buyout_percent": 80,
                        "expected_buyouts": 8,
                        "advertising_spend": 100,
                        "advertising_per_unit": 12.5,
                        "net_profit": 50,
                    }
                ],
            },
            {
                "subject": "Игрушки",
                "store_slug": "tris",
                "store_name": "TRIS",
                "manager": "Менеджер",
                "orders_count": 5,
                "orders_amount": 500,
                "cancel_count": 0,
                "cancel_amount": 0,
                "net_orders_count": 5,
                "net_orders_amount": 500,
                "buyout_percent": 60,
                "buyout_orders_count": 5,
                "expected_buyout_amount": 300,
                "stock": 2,
                "stock_fbs": 0,
                "stock_fbo": 0,
                "stock_fulfillment": 2,
                "stock_days": 1,
                "stock_average_daily_orders": 2,
                "impressions": 100,
                "clicks": 20,
                "advertising_spend": 50,
                "margin_orders_count": 3,
                "margin": 100,
                "purchase_value": 200,
                "margin_complete": True,
                "margin_missing_days": [],
                "daily_calculations": [
                    {
                        "date": "2026-08-30",
                        "available": True,
                        "snapshot_available": True,
                        "orders_count": 5,
                        "net_orders_count": 5,
                        "buyout_percent": 60,
                        "expected_buyouts": 3,
                        "advertising_spend": 50,
                        "advertising_per_unit": 16.67,
                        "net_profit": 20,
                    }
                ],
            },
        ]

        categories = unit_economics._unit_profit_category_rows(rows)

        self.assertEqual(len(categories), 1)
        category = categories[0]
        self.assertEqual(category["name"], "Игрушки")
        self.assertEqual(category["product_count"], 2)
        self.assertEqual(category["store_name"], "2 магазинов")
        self.assertEqual(category["orders_count"], 15)
        self.assertEqual(category["net_orders_amount"], 1300)
        self.assertEqual(category["buyout_percent"], 73.33)
        self.assertEqual(category["ctr"], 15)
        self.assertEqual(category["cpc"], 5)
        self.assertEqual(category["drr"], 13.64)
        self.assertEqual(category["margin"], 500)
        self.assertEqual(category["roi"], 50)
        self.assertEqual(category["stock"], 5)
        self.assertEqual(category["stock_fbs"], 1)
        self.assertEqual(category["stock_fbo"], 2)
        self.assertEqual(category["stock_fulfillment"], 2)
        self.assertEqual(category["stock_days"], 1.67)
        self.assertEqual(category["stock_average_daily_orders"], 3)
        totals = unit_economics._unit_profit_report_totals(rows)
        self.assertEqual(totals, unit_economics._unit_profit_report_totals(categories))
        for field in ("stock", "stock_fbs", "stock_fbo", "stock_fulfillment", "stock_days"):
            self.assertEqual(unit_economics._unit_profit_report_totals([])[field], 0)
        daily = category["daily_calculations"][0]
        self.assertEqual(daily["advertising_spend"], 150)
        self.assertEqual(daily["buyout_percent"], 73.33)
        self.assertEqual(daily["advertising_per_unit"], 13.64)
        self.assertEqual(daily["net_profit"], 41.82)

    def test_unit_profit_totals_keep_available_margin_when_some_rows_are_missing(self) -> None:
        totals = unit_economics._unit_profit_report_totals(
            [
                {
                    "margin": -100,
                    "purchase_value": 200,
                    "margin_complete": True,
                    "margin_missing_days": [],
                },
                {
                    "margin": None,
                    "purchase_value": None,
                    "margin_complete": False,
                    "margin_missing_days": ["2026-08-29"],
                },
                {
                    "margin": 50,
                    "purchase_value": 100,
                    "margin_complete": True,
                    "margin_missing_days": [],
                },
            ]
        )

        self.assertEqual(totals["margin"], -50)
        self.assertEqual(totals["purchase_value"], 300)
        self.assertEqual(totals["roi"], -16.67)
        self.assertFalse(totals["margin_complete"])
        self.assertEqual(totals["margin_missing_days"], ["2026-08-29"])

    def test_unit_profit_report_exposes_daily_margin_inputs(self) -> None:
        inputs = {
            "retail_price": 1000,
            "customer_price": 800,
            "customer_price_with_spp": 800,
            "acquiring_percent": 2,
            "delivery_wb_rub": 50,
            "return_cost_rub": 25,
            "paid_acceptance_cost": 0,
            "storage_wb_rub": 1,
            "turnover_days": 10,
            "commission_percent": 20,
            "purchase_price": 300,
            "fulfillment_cost": 20,
            "team_commission_percent": 5,
            "vat_percent": 20,
            "usn_percent": 6,
            "osno_percent": 0,
            "tax_system": "usn",
            "buyout_percent": 80,
        }

        result = unit_economics._report_daily_calculations(
            date_from=date(2026, 8, 30),
            date_to=date(2026, 8, 30),
            daily_orders={
                "2026-08-30": {
                    "orders_count": 10,
                    "cancel_count": 2,
                    "buyout_percent": 80,
                }
            },
            margin_snapshots={
                "2026-08-30": {
                    "inputs_json": json.dumps(inputs),
                    "unit_margin": 0,
                    "purchase_price": 300,
                }
            },
            live_day=date(2026, 8, 31),
            live_snapshot=None,
            daily_advertising={"2026-08-30": 800},
            fallback_buyout_percent=75,
        )

        self.assertEqual(len(result), 1)
        daily = result[0]
        self.assertTrue(daily["available"])
        self.assertEqual(daily["orders_count"], 10)
        self.assertEqual(daily["net_orders_count"], 8)
        self.assertEqual(daily["advertising_spend"], 800)
        self.assertEqual(daily["advertising_per_unit"], 100)
        self.assertEqual(daily["logistics"], 65)
        self.assertEqual(daily["storage"], 10)
        self.assertEqual(daily["net_revenue"], 605)
        self.assertEqual(daily["net_profit"], 61.67)
        self.assertEqual(daily["vat_value"], 133.33)
        self.assertEqual(daily["usn_value"], 40)

    def test_unit_economics_1c_uses_abc_turnover_and_total_wb_percent(self) -> None:
        product = unit_economics._unit_economics_1c_mock_product(
            "trusthome",
            {"article": "367080326", "name": "Смеситель"},
            {"retail_price": 1232.84, "customer_price_with_spp": 986},
            product_settings=UnitEconomics1CProductSettings(
                store_slug="trusthome",
                article="367080326",
                storage_wb_rub=1.7,
            ),
            product_reference={
                "abc_code": "A",
                "turnover_days": 30,
                "category": "Смесители",
                "subject_commission_percent": 24.62,
                "purchase_price": 901.35,
                "fulfillment_cost": 100.53,
                "team_commission_percent": 3.85,
                "tag_raw": "1 / 0 / SHORT1 / W34 2026",
                "goal_week": 1,
                "goal_day": 0,
                "stock_status": "SHORT1",
                "stock_end_week": "W34 2026",
                "fact_sales": 4,
                "plan_sales": 7,
            },
            wb_extra_tariff_percent=2.3,
        )

        self.assertEqual(product["tag_data"]["code"], "A")
        self.assertEqual(product["tag_data"]["goal_week"], 1)
        self.assertEqual(product["tag_data"]["status"], "SHORT1")
        self.assertEqual(product["tag_data"]["ends"], "W34 2026")
        self.assertEqual(product["tag_data"]["fact"], 4)
        self.assertEqual(product["tag_data"]["plan"], 7)
        self.assertEqual(product["details"]["purchase_cost"], 901.35)
        self.assertEqual(product["details"]["fulfillment_cost"], 100.53)
        self.assertEqual(product["details"]["team_commission_percent"], 3.85)
        self.assertEqual(product["details"]["storage_days"], 30)
        self.assertEqual(product["details"]["storage_sum"], 51)
        self.assertEqual(product["details"]["subject"], "Смесители")
        self.assertEqual(product["details"]["subject_commission_percent"], 24.62)
        self.assertEqual(product["details"]["wb_extra_tariff_percent"], 2.3)
        self.assertEqual(product["details"]["commission_percent"], 26.92)
        self.assertEqual(product["details"]["commission_value"], 331.88)

    def test_unit_economics_1c_price_sync_endpoint(self) -> None:
        with mock.patch.object(
            unit_economics.unit_economics_1c_prices,
            "sync_stores",
            return_value={"rimili": {"ok": True, "rows": 2}},
        ) as sync:
            response = self.client.post(
                "/api/unit-economics-1c/prices/sync",
                headers={"X-Requested-With": "fetch"},
            )
        self.assertEqual(response.status_code, 200, response.text)
        self.assertTrue(response.json()["ok"])
        sync.assert_called_once()

    def test_unit_economics_1c_price_submit_endpoint(self) -> None:
        db.replace_catalog(
            "rimili",
            "WB",
            [{"article": "949558341", "barcode": "2050292584830", "name": "Товар"}],
            NOW,
        )
        report = {
            "ok": True,
            "sent": 1,
            "upload_id": 12345,
            "accepted": [
                {
                    "product_id": "rimili:949558341",
                    "article": "949558341",
                    "base_price": 1000,
                    "target_price": 800,
                    "discount": 20,
                    "calculated_price": 800,
                }
            ],
            "errors": [],
        }
        finalized_report = {
            **report,
            "price_data_refreshed": True,
            "sync": {"ok": True, "rows": 1},
        }
        started = Event()
        release = Event()

        def delayed_submit(store_slug, changes):
            started.set()
            release.wait(timeout=2)
            return report

        with (
            mock.patch.object(
                unit_economics.unit_economics_1c_prices,
                "submit_price_changes",
                side_effect=delayed_submit,
            ) as submit,
            mock.patch.object(
                unit_economics.unit_economics_1c_prices,
                "finalize_price_change_report",
                return_value=finalized_report,
            ) as finalize,
        ):
            response = self.client.post(
                "/api/unit-economics-1c/prices",
                json={
                    "data": [
                        {
                            "store_slug": "rimili",
                            "article": "949558341",
                            "target_price": 800,
                        }
                    ]
                },
                headers={"X-Requested-With": "fetch"},
            )
            self.assertEqual(response.status_code, 202, response.text)
            self.assertTrue(response.json()["ok"])
            job_id = response.json()["job_id"]
            self.assertTrue(started.wait(timeout=1))
            running = self.client.get(f"/api/unit-economics-1c/prices/jobs/{job_id}")
            self.assertIn(running.json()["status"], {"queued", "running"})
            release.set()
            for _ in range(100):
                job_response = self.client.get(f"/api/unit-economics-1c/prices/jobs/{job_id}")
                if job_response.json()["status"] not in {"queued", "running"}:
                    break
                time.sleep(0.01)
            self.assertEqual(job_response.status_code, 200, job_response.text)
            self.assertEqual(job_response.json()["status"], "success")
            result = job_response.json()["result"]
            self.assertEqual(result["accepted_product_ids"], ["rimili:949558341"])
        submit.assert_called_once_with(
            "rimili",
            [
                {
                    "article": "949558341",
                    "target_price": 800.0,
                    "target_kind": "retail",
                }
            ],
        )
        finalize.assert_called_once_with("rimili", report)
        self.assertTrue(result["price_data_refreshed"])

    def test_unit_profit_report_stock_matches_wb_unit_and_export(self) -> None:
        from io import BytesIO

        import openpyxl

        db.upsert_mp_stock("rimili", "949558341", "WB", "fbo", 2, NOW)
        params = {"store": "rimili", "article": "949558341",
                  "date_from": "2026-08-01", "date_to": "2026-08-31"}
        stock_fields = {
            "stock": ("total", "Всего, шт."),
            "stock_fbs": ("fbs", "FBS, шт."),
            "stock_fbo": ("fbo", "FBO, шт."),
            "stock_fulfillment": ("fulfillment", "ФФ, шт."),
            "stock_days": ("days", "Хватит, дней"),
        }
        # A single order also catches premature rounding of average daily demand.
        for orders_count in (0, 1, 42):
            with self.subTest(orders_count=orders_count), mock.patch.object(
                unit_economics.unit_economics_1c,
                "load_product_average_daily_orders",
                return_value={
                    ("rimili", "949558341"): {"orders_count": orders_count, "period_days": 21},
                },
            ) as demand:
                product = self._unit_economics_product("rimili", "949558341")
                response = self.client.get("/api/unit-economics-1c/reports/unit-profit", params=params)
                categories = self.client.get(
                    "/api/unit-economics-1c/reports/unit-profit",
                    params={**params, "group_by": "subject"},
                )
                download = self.client.get(
                    "/sales/unit-economics-1c/reports/unit-profit.xlsx", params=params,
                )
                for result in (response, categories, download):
                    self.assertEqual(result.status_code, 200)
                payload = response.json()
                self.assertEqual(len(payload["rows"]), 1)
                row = payload["rows"][0]
                self.assertEqual((row["stock_fbs"], row["stock_fbo"], row["stock_fulfillment"]), (3, 2, 5))
                self.assertEqual(row["stock"], 10)
                self.assertEqual(row["stock_average_daily_orders"], orders_count / 21)
                self.assertEqual(row["stock_days"], round(210 / orders_count, 2) if orders_count else 0)
                for field, (source_key, _) in stock_fields.items():
                    self.assertEqual(row[field], product["stock"][source_key])
                    self.assertEqual(payload["totals"][field], row[field])
                    self.assertEqual(categories.json()["rows"][0][field], row[field])
                # A past report uses the same current stock demand as the WB unit page.
                for call in demand.call_args_list:
                    self.assertEqual(call.kwargs, {"period_days": 21})

                workbook = openpyxl.load_workbook(BytesIO(download.content))
                sheet = workbook["Расчёт маржи и ROI"]
                headers = {cell.value: cell.column for cell in sheet[5]}
                self.assertIn("U4:Y4", {str(cell_range) for cell_range in sheet.merged_cells.ranges})
                self.assertEqual(sheet["U4"].value, "Остатки")
                self.assertEqual(sheet["Z4"].value, "Реклама")
                self.assertEqual(sheet["BP4"].value, "Результат")
                for field, (_, label) in stock_fields.items():
                    self.assertEqual(sheet.cell(6, headers[label]).value, row[field])
                    self.assertEqual(sheet.cell(sheet.max_row, headers[label]).value, row[field])
                self.assertEqual(sheet.cell(6, headers["Хватит, дней"]).number_format, "#,##0.00")
                workbook.close()

    def test_unit_profit_report_supports_period_filters_and_totals(self) -> None:
        wb_funnel_orders._replace_day(
            "rimili",
            datetime(2026, 8, 13).date(),
            [("949558341", "R-1", "Товар", 4, 2_000, 1, 500, 3, 1_500, 80)],
        )
        wb_funnel_orders._replace_day(
            "rimili",
            datetime(2026, 8, 19).date(),
            [("949558341", "R-1", "Товар", 5, 2_500, 1, 500, 4, 2_000, 80)],
        )
        wb_funnel_orders._replace_day(
            "rimili",
            datetime(2026, 8, 20).date(),
            [("949558341", "R-1", "Товар", 7, 3_500)],
        )
        metrics = {
            ("rimili", "949558341"): {
                "period_from": "2026-08-13",
                "period_to": "2026-08-19",
                "period_days": 7,
                "spend": 210,
                "average_daily_spend": 30,
                "impressions": 1000,
                "clicks": 25,
                "ctr": 2.5,
                "cpc": 8.4,
                "orders_amount": 5000,
                "orders_count": 10,
                "cancel_amount": 1000,
                "cancel_count": 2,
                "net_orders_amount": 4000,
                "net_orders_count": 8,
                "funnel_updated_at": "2026-08-19T08:00:00+00:00",
                "funnel_source_version": 3,
                "funnel_vendor_code": "R-1",
                "funnel_product_name": "Товар WB",
                "sold_count": 8,
                "buyout_percent": 80,
                "range_buyout_percent": 80,
                "buyout_count": 7,
                "buyout_amount": 3500,
                "buyout_percent_weight": 9,
                "buyout_orders_count": 10,
                "buyout_orders_amount": 5000,
                "buyout_cancel_count": 2,
                "buyout_cancel_amount": 1000,
                "buyout_net_orders_count": 8,
                "buyout_net_orders_amount": 4000,
                "buyout_period_from": "2026-08-13",
                "buyout_period_to": "2026-08-19",
                "buyout_updated_at": "2026-08-19T08:00:00+00:00",
                "buyout_source_version": 2,
                "drr": 4.2,
                "daily": [
                    {
                        "date": "2026-08-13",
                        "advertising_spend": 90,
                        "orders_count": 4,
                        "orders_amount": 2000,
                        "cancel_count": 1,
                        "cancel_amount": 500,
                        "buyout_percent": 80,
                    },
                    {
                        "date": "2026-08-19",
                        "advertising_spend": 120,
                        "orders_count": 5,
                        "orders_amount": 2500,
                        "cancel_count": 1,
                        "cancel_amount": 500,
                        "buyout_percent": 80,
                    },
                ],
            }
        }
        filters_response = self.client.get(
            "/api/unit-economics-1c/reports/unit-profit/filters",
            params={"store": "rimili"},
        )
        with mock.patch.object(
            unit_economics.unit_economics_1c,
            "load_product_metrics",
            return_value=metrics,
        ):
            response = self.client.get(
                "/api/unit-economics-1c/reports/unit-profit",
                params={
                    "date_from": "2026-08-13",
                    "date_to": "2026-08-19",
                    "store": "rimili",
                },
            )
            download = self.client.get(
                "/sales/unit-economics-1c/reports/unit-profit.xlsx",
                params={
                    "date_from": "2026-08-13",
                    "date_to": "2026-08-19",
                    "store": "rimili",
                },
            )
            category_download = self.client.get(
                "/sales/unit-economics-1c/reports/unit-profit.xlsx",
                params={
                    "date_from": "2026-08-13",
                    "date_to": "2026-08-19",
                    "store": "rimili",
                    "group_by": "subject",
                    "daily_details": "1",
                },
            )
            category_response = self.client.get(
                "/api/unit-economics-1c/reports/unit-profit",
                params={
                    "date_from": "2026-08-13",
                    "date_to": "2026-08-19",
                    "store": "rimili",
                    "group_by": "subject",
                    "daily_details": "1",
                    "page": "1",
                    "page_size": "25",
                },
            )
            filtered_daily_download = self.client.get(
                "/sales/unit-economics-1c/reports/unit-profit.xlsx",
                params={
                    "date_from": "2026-08-13",
                    "date_to": "2026-08-19",
                    "store": "rimili",
                    "article": "949558341",
                    "daily_details": "1",
                },
            )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(filters_response.status_code, 200, filters_response.text)
        self.assertTrue(filters_response.json()["ok"])
        self.assertIn("949558341", {
            item["article"] for item in filters_response.json()["filters"]["articles"]
        })
        payload = response.json()
        self.assertEqual((payload["period_from"], payload["period_to"]), ("2026-08-13", "2026-08-19"))
        row = next(item for item in payload["rows"] if item["article"] == "949558341")
        self.assertEqual(row["orders_count"], 9)
        self.assertNotIn("funnel_orders_count", row)
        self.assertEqual(row["orders_amount"], 4500)
        self.assertEqual(
            (row["cancel_count"], row["cancel_amount"], row["net_orders_count"], row["net_orders_amount"]),
            (2, 1000, 7, 3500),
        )
        self.assertEqual(row["buyout_percent"], 80)
        self.assertEqual(row["ctr"], 2.5)
        self.assertEqual(row["cpc"], 8.4)
        self.assertEqual(row["advertising_spend"], 210)
        self.assertNotIn("legal_entity", row)
        self.assertIn("subject", row)
        orders_amounts = [item["orders_amount"] for item in payload["rows"]]
        self.assertEqual(orders_amounts, sorted(orders_amounts, reverse=True))
        self.assertEqual(payload["totals"]["orders_count"], 9)
        self.assertEqual(payload["totals"]["cancel_count"], 2)
        self.assertEqual(payload["totals"]["net_orders_count"], 7)
        self.assertEqual(payload["totals"]["cancel_amount"], 1000)
        self.assertEqual(payload["totals"]["net_orders_amount"], 3500)
        self.assertEqual(payload["totals"]["buyout_percent"], 80)
        self.assertEqual(payload["totals"]["ctr"], 2.5)
        self.assertNotIn("legal_entities", payload["filters"])
        self.assertIn("subjects", payload["filters"])
        self.assertIn("managers", payload["filters"])
        self.assertIn("articles", payload["filters"])
        self.assertEqual(payload["funnel_daily_rows"], [])
        self.assertEqual(payload["funnel_weekly_rows"], [])
        self.assertEqual(payload["pagination"]["page"], 1)
        self.assertEqual(payload["pagination"]["page_size"], 50)
        self.assertEqual(payload["pagination"]["total_count"], len(payload["rows"]))
        self.assertEqual(payload["pagination"]["total_pages"], 1)
        self.assertFalse(payload["pagination"]["enabled"])
        self.assertTrue(
            {
                "retail_price",
                "customer_price",
                "acquiring_percent",
                "acquiring_value",
                "delivery_with_returns",
                "advertising_per_unit",
                "purchase_cost",
                "tax_value",
                "net_revenue",
                "margin_orders_count",
                "purchase_value",
                "roi",
            }.issubset(row)
        )
        self.assertNotIn("unit_margin", row)

        self.assertEqual(download.status_code, 200, download.text)
        self.assertEqual(
            download.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(
            "unit_profit_2026-08-13_2026-08-19.xlsx",
            download.headers["content-disposition"],
        )
        from io import BytesIO

        import openpyxl

        workbook = openpyxl.load_workbook(BytesIO(download.content), data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            ["Расчёт маржи и ROI", "Воронка по дням", "Воронка 7 дней"],
        )
        sheet = workbook["Расчёт маржи и ROI"]
        self.assertEqual(sheet["A1"].value, "Отчёт по юниточной прибыли — полный расчёт")
        headers = {
            sheet.cell(row=5, column=column).value: column
            for column in range(1, sheet.max_column + 1)
        }
        for title in (
            "Всего заказов, шт.",
            "Всего отмен, шт.",
            "Заказы − отмены, шт.",
            "Всего заказов, ₽",
            "Всего отмен, ₽",
            "ТО после отмен, ₽",
            "Выкупы, шт.",
            "Сумма выкупов, ₽",
            "Процент выкупа",
            "Реклама на выкупленную ед., ₽",
            "Выкупы в расчёте маржи",
            "Маржа периода, ₽",
            "Закупка периода, ₽",
            "ROI",
        ):
            self.assertIn(title, headers)
        self.assertNotIn("Маржа на штуку, ₽", headers)
        product_row = next(
            row_index
            for row_index in range(6, sheet.max_row)
            if sheet.cell(row=row_index, column=2).value == "949558341"
        )
        self.assertEqual(sheet.cell(row=product_row, column=headers["Всего заказов, шт."]).value, 9)
        self.assertEqual(sheet.cell(row=product_row, column=headers["Всего отмен, шт."]).value, 2)
        self.assertEqual(sheet.cell(row=product_row, column=headers["Заказы − отмены, шт."]).value, 7)
        self.assertEqual(sheet.cell(row=product_row, column=headers["Всего заказов, ₽"]).value, 4500)
        self.assertEqual(sheet.cell(row=product_row, column=headers["Всего отмен, ₽"]).value, 1000)
        self.assertEqual(sheet.cell(row=product_row, column=headers["ТО после отмен, ₽"]).value, 3500)
        self.assertEqual(sheet.cell(row=sheet.max_row, column=1).value, "ИТОГО")
        self.assertEqual(sheet.freeze_panes, "F6")
        self.assertTrue(sheet.auto_filter.ref.startswith("A5:"))

        self.assertEqual(category_download.status_code, 200, category_download.text)
        self.assertEqual(category_response.status_code, 200, category_response.text)
        category_payload = category_response.json()
        self.assertTrue(category_payload["rows"])
        self.assertTrue(all(item["row_kind"] == "category" for item in category_payload["rows"]))
        self.assertFalse(category_payload["daily_details"])
        self.assertFalse(category_payload["rows"][0]["daily_calculations"])
        self.assertFalse(category_payload["pagination"]["enabled"])
        self.assertEqual(category_payload["pagination"]["page_size"], 25)
        category_workbook = openpyxl.load_workbook(
            BytesIO(category_download.content),
            data_only=False,
        )
        category_sheet = category_workbook["Расчёт маржи и ROI"]
        self.assertEqual(
            category_sheet["A1"].value,
            "Отчёт по юниточной прибыли — по категориям",
        )
        self.assertEqual(category_sheet["G2"].value, 1)
        self.assertEqual(category_sheet.max_column, 71)

        self.assertEqual(filtered_daily_download.status_code, 200, filtered_daily_download.text)
        filtered_daily_workbook = openpyxl.load_workbook(
            BytesIO(filtered_daily_download.content),
            data_only=False,
        )
        filtered_daily_sheet = filtered_daily_workbook["Расчёт маржи и ROI"]
        self.assertEqual(filtered_daily_sheet["G2"].value, 1)
        self.assertEqual(filtered_daily_sheet.max_column, 71 + 7 * 20)
        self.assertEqual(
            filtered_daily_sheet.cell(row=4, column=72).value.date(),
            date(2026, 8, 13),
        )
        self.assertEqual(
            sum(
                filtered_daily_sheet.cell(row=5, column=column).value
                == "Расходы на рекламу, ₽"
                for column in range(72, filtered_daily_sheet.max_column + 1)
            ),
            7,
        )
        filtered_articles = [
            filtered_daily_sheet.cell(row=row_index, column=2).value
            for row_index in range(6, filtered_daily_sheet.max_row)
        ]
        self.assertEqual(filtered_articles, ["949558341"])
        daily_sheet = workbook["Воронка по дням"]
        daily_headers = [
            daily_sheet.cell(row=4, column=column).value
            for column in range(1, daily_sheet.max_column + 1)
        ]
        self.assertIn("Всего отмен, шт.", daily_headers)
        self.assertIn("ТО после отмен, ₽", daily_headers)
        self.assertIn("Выкупы, шт.", daily_headers)
        self.assertIn("Сумма выкупов, ₽", daily_headers)

        page = self.client.get("/sales/unit-economics-1c/reports/unit-profit")
        self.assertNotIn("Юрлицо", page.text)
        self.assertNotIn('id="ue1cr-legal"', page.text)
        self.assertNotIn('class="ue1cr-page-head"', page.text)
        self.assertNotIn('id="ue1cr-summary"', page.text)
        self.assertNotIn('class="ue1cr-total-row" id="ue1cr-total"', page.text)
        self.assertIn('data-report-total-count', page.text)
        self.assertIn('data-report-total="orders_count"', page.text)
        self.assertIn('data-report-total="cancel_count"', page.text)
        self.assertIn('data-report-total="net_orders_count"', page.text)
        self.assertIn('data-report-total="cancel_amount"', page.text)
        self.assertIn('data-report-total="net_orders_amount"', page.text)
        self.assertNotIn('data-report-total="funnel_orders_count"', page.text)
        self.assertNotIn(">Заказов</span>", page.text)
        self.assertIn("Всего заказов, шт.", page.text)
        self.assertIn("Всего отмен, шт.", page.text)
        self.assertIn("Заказы − отмены, шт.", page.text)
        self.assertIn("Выкуп</span>", page.text)
        self.assertNotIn("Маржа на штуку, ₽", page.text)
        self.assertNotIn('data-filter-column="26"', page.text)
        for field in ("stock", "stock_fbs", "stock_fbo", "stock_fulfillment", "stock_days"):
            self.assertIn(f'data-report-total="{field}"', page.text)
        self.assertNotIn('id="ue1cr-details-dialog"', page.text)
        self.assertNotIn('<tfoot id="ue1cr-total">', page.text)
        self.assertIn('<div class="ue1cr-filter"><span>Менеджеры</span>', page.text)
        self.assertIn('id="ue1cr-manager"', page.text)
        self.assertIn('id="ue1cr-article-search"', page.text)
        self.assertIn('id="ue1cr-show-images"', page.text)
        self.assertIn('id="ue1cr-columns-toggle"', page.text)
        self.assertIn('id="ue1cr-column-list"', page.text)
        self.assertIn('id="ue1cr-export"', page.text)
        self.assertIn('id="ue1cr-submit">Сформировать</button>', page.text)
        self.assertIn('Настройте параметры и нажмите «Сформировать».', page.text)
        self.assertIn('id="ue1cr-pagination"', page.text)
        self.assertIn('id="ue1cr-page-size"', page.text)
        self.assertIn("unit-profit.xlsx", page.text)
        self.assertIn('data-filter-column="7" data-filter-type="number"', page.text)
        report_script = (
            Path(__file__).resolve().parents[2] / "static" / "unit-economics-1c-report.js"
        ).read_text(encoding="utf-8")
        self.assertIn("function turnover(item)", report_script)
        self.assertIn("if (column.format === 'turnover') content = turnover(raw)", report_script)
        self.assertIn("{ key: 'margin', label: 'Маржа периода, ₽', format: 'number'", report_script)
        self.assertIn("data-filter-value", report_script)
        self.assertIn("setHeaderTotal('margin', value(total.margin)", report_script)
        self.assertNotIn("+ rub(row.orders_amount)", report_script)
        self.assertNotIn("row.unit_margin", report_script)
        self.assertIn("query.set('page', String(state.page))", report_script)
        self.assertIn("query.set('page_size', String(state.pageSize))", report_script)
        self.assertIn("state.rows = result.rows || []", report_script)
        self.assertIn("window.localStorage.setItem(preferenceKey", report_script)
        self.assertIn("hiddenColumns", report_script)
        self.assertIn("data-column-metric", report_script)
        self.assertIn("function visibleColumnsForGroup(group)", report_script)
        self.assertIn("margin_undercovered_days", report_script)
        self.assertIn("ue1cr-margin-cell--partial", report_script)
        self.assertIn("Маржа рассчитана только по датам с доступными снимками", report_script)
        report_styles = (
            Path(__file__).resolve().parents[2] / "static" / "unit-economics-1c-report.css"
        ).read_text(encoding="utf-8")
        self.assertIn("td.ue1cr-margin-cell--partial", report_styles)
        self.assertIn("text-decoration-style: dotted", report_styles)
        self.assertIn("ue1cr-page--daily-hidden", report_script)
        self.assertIn("function reportQuery(includeView)", report_script)
        self.assertIn("function updateExportLink()", report_script)
        self.assertIn("async function downloadExcel(event)", report_script)
        self.assertIn("Формируем Excel…", report_script)
        self.assertIn("window.URL.createObjectURL(blob)", report_script)
        self.assertIn("nodes.export.addEventListener('click', downloadExcel)", report_script)
        self.assertIn("var AUTO_REFRESH_INTERVAL_MS = 5 * 60 * 1000", report_script)
        self.assertIn("function refreshReportIfStale(force)", report_script)
        self.assertIn("if (!state.reportLoaded || document.hidden", report_script)
        self.assertIn("async function loadFilterOptions()", report_script)
        self.assertIn("reportLoaded: false", report_script)
        self.assertTrue(report_script.rstrip().endswith("loadFilterOptions();\n})();"))
        self.assertIn("load({ silent: true })", report_script)
        self.assertIn("timeZone: 'Europe/Moscow'", report_script)
        self.assertIn("document.addEventListener('visibilitychange'", report_script)
        self.assertNotIn("data-report-details", report_script)

        invalid = self.client.get(
            "/api/unit-economics-1c/reports/unit-profit",
            params={"date_from": "2026-08-20", "date_to": "2026-08-19"},
        )
        self.assertEqual(invalid.status_code, 422)
        invalid_export = self.client.get(
            "/sales/unit-economics-1c/reports/unit-profit.xlsx",
            params={"date_from": "2026-08-20", "date_to": "2026-08-19"},
        )
        self.assertEqual(invalid_export.status_code, 422)

    def test_unit_profit_report_paginates_only_daily_product_details(self) -> None:
        products = [
            {
                "article": str(100_000_000 + index),
                "name": f"Товар {index:02d}",
                "barcode": "",
                "image_url": "",
                "fbs_stock": 0,
                "fbo_stock": 0,
                "ff_available": 0,
            }
            for index in range(60)
        ]
        with (
            mock.patch.object(unit_economics.db, "get_stock_items", return_value=products),
            mock.patch.object(
                unit_economics.unit_economics_1c,
                "load_product_metrics",
                return_value={},
            ),
            mock.patch.object(
                unit_economics.unit_economics_1c,
                "load_product_average_daily_orders",
                return_value={},
            ),
        ):
            response = self.client.get(
                "/api/unit-economics-1c/reports/unit-profit",
                params={
                    "date_from": "2026-08-01",
                    "date_to": "2026-08-02",
                    "store": "rimili",
                    "page": "2",
                    "page_size": "25",
                },
            )
            daily_response = self.client.get(
                "/api/unit-economics-1c/reports/unit-profit",
                params={
                    "date_from": "2026-08-01",
                    "date_to": "2026-08-02",
                    "store": "rimili",
                    "daily_details": "1",
                    "page": "2",
                    "page_size": "25",
                },
            )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(len(payload["rows"]), 60)
        self.assertEqual(
            payload["pagination"],
            {
                "enabled": False,
                "page": 1,
                "page_size": 25,
                "total_count": 60,
                "total_pages": 1,
            },
        )
        self.assertEqual(daily_response.status_code, 200, daily_response.text)
        daily_payload = daily_response.json()
        self.assertEqual(len(daily_payload["rows"]), 25)
        self.assertTrue(daily_payload["pagination"]["enabled"])
        self.assertEqual(daily_payload["pagination"]["page"], 2)
        self.assertEqual(daily_payload["pagination"]["total_pages"], 3)

    def test_unit_profit_report_filters_by_manager_and_limits_regular_users(self) -> None:
        stock_item = next(
            item
            for item in db.list_unit_economics_1c_active_wb_stock_items()
            if item["store_slug"] == "rimili" and item["article"] == "949558341"
        )
        db.replace_unit_economics_1c_source_values(
            [
                {
                    "stock_item_id": stock_item["id"],
                    "manager": "Андрей Китов",
                    "source_sheet_id": 10,
                    "source_sheet_title": "RIMILI WB",
                    "source_row": 3,
                }
            ],
            {},
            NOW,
        )

        filtered = self.client.get(
            "/api/unit-economics-1c/reports/unit-profit",
            params=[("store", "rimili"), ("manager", "Андрей Китов"), ("article", "949558341")],
        )
        self.assertEqual(filtered.status_code, 200, filtered.text)
        self.assertEqual([row["article"] for row in filtered.json()["rows"]], ["949558341"])
        self.assertEqual(filtered.json()["rows"][0]["manager"], "Андрей Китов")

        self.app.state.container.identity.user_for_token.return_value = {
            **self.user,
            "role": "admin",
        }
        admin_page = self.client.get("/sales/unit-economics-1c/reports/unit-profit")
        self.assertIn('<div class="ue1cr-filter"><span>Менеджеры</span>', admin_page.text)

        self.app.state.container.identity.user_for_token.return_value = {
            **self.user,
            "id": 4,
            "full_name": "Андрей Китов",
            "login": "Kitov",
            "role": "user",
            "store_slugs": ["rimili"],
        }
        own = self.client.get("/api/unit-economics-1c/reports/unit-profit", params={"store": "rimili"})
        self.assertEqual([row["article"] for row in own.json()["rows"]], ["949558341"])
        self.assertTrue(own.json()["manager_scope"]["matched"])
        manager_page = self.client.get("/sales/unit-economics-1c/reports/unit-profit")
        self.assertNotIn('id="ue1cr-manager"', manager_page.text)
        self.assertNotIn('id="ue1cr-manager-summary"', manager_page.text)
        self.assertNotIn('id="ue1cr-manager-options"', manager_page.text)
        self.assertNotIn("<span>Менеджеры</span>", manager_page.text)
        report_script = (
            Path(__file__).resolve().parents[2] / "static" / "unit-economics-1c-report.js"
        ).read_text(encoding="utf-8")
        self.assertIn("if (!summaryNodes[kind]) return;", report_script)
        self.assertIn("if (!optionNodes[kind]) return;", report_script)

        self.app.state.container.identity.user_for_token.return_value = {
            **self.user,
            "id": 5,
            "full_name": "Антон Ефимов",
            "login": "Efimov",
            "role": "user",
            "store_slugs": ["rimili"],
        }
        unrelated = self.client.get(
            "/api/unit-economics-1c/reports/unit-profit",
            params={"store": "rimili"},
        )
        self.assertEqual(unrelated.status_code, 200, unrelated.text)
        self.assertEqual(unrelated.json()["rows"], [])
        self.assertFalse(unrelated.json()["manager_scope"]["matched"])

        self.app.state.container.identity.user_for_token.return_value = {
            **self.user,
            "id": 6,
            "full_name": "Андрей Китов Старший",
            "login": "KitovSenior",
            "role": "user",
            "store_slugs": ["rimili"],
        }
        partial_name = self.client.get(
            "/api/unit-economics-1c/reports/unit-profit",
            params={"store": "rimili"},
        )
        self.assertEqual(partial_name.status_code, 200, partial_name.text)
        self.assertEqual(partial_name.json()["rows"], [])
        self.assertFalse(partial_name.json()["manager_scope"]["matched"])

        self.app.state.container.identity.user_for_token.return_value = {
            **self.user,
            "id": 4,
            "full_name": "Андрей Китов",
            "login": "Kitov",
            "role": "user",
            "store_slugs": ["rimili"],
        }
        with mock.patch.object(
            unit_economics.db,
            "get_unit_economics_1c_product_reference_rows",
            return_value=[],
        ):
            no_assignments = self.client.get(
                "/api/unit-economics-1c/reports/unit-profit",
                params={"store": "rimili"},
            )
        self.assertEqual(no_assignments.status_code, 200, no_assignments.text)
        self.assertEqual(no_assignments.json()["rows"], [])
        self.assertTrue(no_assignments.json()["manager_scope"]["restricted"])
        self.assertFalse(no_assignments.json()["manager_scope"]["matched"])

    def test_unit_economics_column_preferences_are_personal_and_server_backed(self) -> None:
        conn = core.get_connection()
        for user_id in (1, 2):
            conn.execute(
                """
                INSERT OR IGNORE INTO users
                    (id, full_name, google_email, login, password_hash, role, is_active,
                     can_edit_stock, can_manage_users, created_at)
                VALUES (?, ?, '', ?, 'hash', 'user', 1, 1, 0, ?)
                """,
                (user_id, f"User {user_id}", f"user-{user_id}", NOW),
            )
        conn.commit()
        conn.close()
        payload = {
            "order": ["product", "stock", "comments", "current"],
            "hidden": ["tag", "product"],
        }
        saved = self.client.put(
            "/api/unit-economics-1c/preferences/columns",
            json=payload,
            headers={"X-Requested-With": "fetch"},
        )
        self.assertEqual(saved.status_code, 200, saved.text)
        self.assertEqual(saved.json()["preferences"]["order"][:4], payload["order"])
        self.assertEqual(saved.json()["preferences"]["hidden"], ["tag"])
        db.save_ui_preference(2, "unit_economics_1c.columns", {"order": ["product"], "hidden": []})
        self.assertEqual(db.get_ui_preference(1, "unit_economics_1c.columns")["hidden"], ["tag"])
        self.assertEqual(db.get_ui_preference(2, "unit_economics_1c.columns")["hidden"], [])
        page = self.client.get("/sales/unit-economics-1c")
        self.assertIn('"columnPreferences": {"order": ["product", "stock", "comments", "current"', page.text)

    def test_unit_economics_chart_and_column_drag_controls_are_rendered(self) -> None:
        root = Path(__file__).resolve().parents[2]
        script = (root / "static" / "unit-economics-1c.js").read_text(encoding="utf-8")
        styles = (root / "static" / "unit-economics-1c.css").read_text(encoding="utf-8")
        template = (root / "templates" / "unit_economics_1c_content.html").read_text(encoding="utf-8")
        self.assertIn('data-chart-series="orders"', template)
        self.assertIn('data-chart-series="stock"', template)
        self.assertIn("data-chart-compare", template)
        self.assertIn('id="ue1c-chart-daily-sales"', template)
        self.assertIn("Заказы воронки ×10", template)
        self.assertIn("Воронка неделей ранее", template)
        self.assertIn("var orderScale = 10", script)
        self.assertIn("var history = allHistory.slice(-14)", script)
        self.assertIn("var orderBaselineY = usesMoneyAxis ? moneyY(0) : bottom", script)
        self.assertIn("data-chart-line-series", script)
        self.assertIn("showChartSeriesTooltip", script)
        self.assertIn("renderChartDailySales", script)
        self.assertNotIn('class="ue1c-chart-bar-value', script)
        self.assertIn("<table><tbody>", script)
        self.assertIn('<th scope="row">Заказы воронки</th>', script)
        self.assertNotIn('Заказы / выкуп</th>', script)
        self.assertNotIn("line(previousHistory", script)
        self.assertIn(".ue1c-chart-tooltip td", styles)
        self.assertIn("width: 620px", styles)
        self.assertIn("height: 260px", styles)
        self.assertIn(".ue1c-chart-daily-sales-value", styles)
        self.assertIn("finite(item.orders_count, 0)", script)
        self.assertIn("line(history, 'stock_units'", script)
        self.assertIn("(group.fixed ? 'false' : 'true')", script)
        self.assertIn("addEventListener('drop'", script)
        self.assertIn("/api/unit-economics-1c/preferences/columns", script)

    def test_unit_economics_1c_price_preview_endpoint_is_point_only(self) -> None:
        db.replace_catalog(
            "rimili",
            "WB",
            [{"article": "949558341", "barcode": "2050292584830", "name": "Товар"}],
            NOW,
        )
        report = {
            "ok": True,
            "planned": 1,
            "accepted": [
                {
                    "product_id": "rimili:949558341",
                    "article": "949558341",
                    "target_kind": "spp",
                    "target_price": 700,
                    "base_price": 2002,
                    "discount": 50,
                    "predicted_spp_price": 700,
                }
            ],
            "errors": [],
        }
        with mock.patch.object(
            unit_economics.unit_economics_1c_prices,
            "preview_price_changes",
            return_value=report,
        ) as preview:
            response = self.client.post(
                "/api/unit-economics-1c/prices/preview",
                json={
                    "data": [
                        {
                            "store_slug": "rimili",
                            "article": "949558341",
                            "target_kind": "spp",
                            "target_price": 700,
                        }
                    ]
                },
                headers={"X-Requested-With": "fetch"},
            )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["accepted"][0]["predicted_spp_price"], 700)
        preview.assert_called_once_with(
            "rimili",
            [
                {
                    "article": "949558341",
                    "target_price": 700.0,
                    "target_kind": "spp",
                }
            ],
        )

        too_many = self.client.post(
            "/api/unit-economics-1c/prices/preview",
            json={
                "data": [
                    {
                        "store_slug": "rimili",
                        "article": "949558341",
                        "target_price": 700,
                    },
                    {
                        "store_slug": "rimili",
                        "article": "949558341",
                        "target_price": 701,
                    },
                ]
            },
            headers={"X-Requested-With": "fetch"},
        )
        self.assertEqual(too_many.status_code, 422)

    def test_unit_economics_1c_product_settings_save_and_render(self) -> None:
        wb_funnel_orders._replace_product_metrics(
            "rimili",
            datetime(2026, 8, 20).date(),
            datetime(2026, 8, 26).date(),
            [("949558341", 10, 10_000, 80)],
        )
        payload = {
            "article": "949558341",
            "delivery_wb_rub": 100,
            "return_cost_rub": 50,
            "volume_l": 1.2,
            "storage_wb_rub": 3,
        }
        response = self.client.put(
            "/api/unit-economics-1c/product-settings/rimili",
            json=payload,
            headers={"X-Requested-With": "fetch"},
        )
        self.assertEqual(response.status_code, 200, response.text)
        saved = response.json()["settings"]
        self.assertEqual(saved["paid_acceptance_cost"], 0)
        self.assertEqual(saved["delivery_with_returns"], 130)

        product = self._unit_economics_product("rimili", "949558341")
        self.assertEqual(product["details"]["delivery_wb_rub"], 100)
        self.assertEqual(product["details"]["buyout_percent"], 80)
        self.assertEqual(product["details"]["delivery_with_returns"], 130)
        self.assertIsNone(product["details"]["subject"])
        script = (Path(__file__).resolve().parents[2] / "static" / "unit-economics-1c.js").read_text(
            encoding="utf-8"
        )
        self.assertIn("Выкуп WB · период кабинета", script)
        self.assertIn("Выкуп · значение кабинета", script)
        self.assertNotIn("Выкуп для логистики", script)
        self.assertNotIn("editableParameter('Выкуп", script)
        self.assertNotIn("data-product-setting=\"buyout_percent\"", script)

        invalid = self.client.put(
            "/api/unit-economics-1c/product-settings/rimili",
            json={**payload, "buyout_percent": 80},
            headers={"X-Requested-With": "fetch"},
        )
        self.assertEqual(invalid.status_code, 422)

    def test_unit_economics_1c_does_not_render_price_only_history(self) -> None:
        db.upsert_unit_economics_1c_daily_prices(
            [
                {
                    "store_slug": "rimili",
                    "article": "STALE-HISTORY-ONLY",
                    "day": "2026-08-19",
                    "marketplace": "WB",
                    "nm_id": "999999999",
                    "currency": "RUB",
                    "customer_price_orders_count": 0,
                    "updated_at": NOW,
                }
            ]
        )
        articles = {item["article"] for item in self._unit_economics_products()}
        self.assertNotIn("STALE-HISTORY-ONLY", articles)

    def test_yandex_stock_shows_combined_fby_and_fbs(self) -> None:
        db.replace_catalog(
            "rimili",
            "YANDEX MARKET",
            [{"article": "YA-1", "barcode": "123", "name": "Яндекс товар"}],
            NOW,
        )
        db.upsert_mp_stock("rimili", "YA-1", "YANDEX MARKET", "fbo", 19, NOW)
        db.upsert_mp_stock("rimili", "YA-1", "YANDEX MARKET", "fbs", 17, NOW)
        db.replace_mp_warehouse_stock(
            "rimili",
            "YANDEX MARKET",
            "fbs_149217490",
            [("YA-1", "Afflatus", None, 17, NOW)],
        )

        page = self.client.get("/stock/rimili", params={"mp": "YANDEX MARKET"})
        self.assertEqual(page.status_code, 200, page.text[:500])
        self.assertIn("FBY — склады Маркета", page.text)
        self.assertIn("FBS — склады продавца", page.text)
        self.assertIn('tot-grand">36</strong>', page.text)
        self.assertIn('tot-fbo">19</strong>', page.text)
        self.assertIn('tot-fbs">17</strong>', page.text)

        fbs = self.client.get("/stock/rimili/fbs", params={"mp": "YANDEX MARKET"})
        self.assertEqual(fbs.json(), {"fbs": {"YA-1": 17}})

        detail = self.client.get(
            "/stock/rimili/article-detail",
            params={"mp": "YANDEX MARKET", "article": "YA-1"},
        )
        fulfillment = next(row for row in detail.json()["warehouses"] if row["name"] == "AFFLATUS Купавна")
        self.assertEqual(fulfillment["fbs"], 17)

        warehouses = self.client.get("/stock/rimili/warehouses", params={"mp": "YANDEX MARKET"})
        self.assertEqual(warehouses.status_code, 200, warehouses.text[:500])
        self.assertIn("Afflatus", warehouses.text)

    def test_page_and_download_errors_are_reported(self) -> None:
        cases = (
            ("/stock/unknown", 404),
            ("/stock/unknown/fbs", 404),
            ("/stock/unknown/warehouses", 404),
            ("/stock/unknown/operations", 404),
            ("/stock-2/details/unknown", 404),
            ("/admin/operations/999999/xlsx", 404),
            ("/stock/rimili/article-detail?mp=WB&article=missing", 200),
        )
        for path, status in cases:
            with self.subTest(path=path):
                self.assertEqual(self.client.get(path).status_code, status)

    def test_sales_decision_and_rnp_api_contracts(self) -> None:
        dashboard = {
            "ok": True,
            "marketplace": "WB",
            "period": {},
            "summary": {},
            "opportunities": [],
            "products": [],
            "sync": [],
        }
        with mock.patch.object(decision_routes.decision_service, "dashboard", return_value=dashboard):
            response = self.client.get("/api/decision-center")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])

        with mock.patch.object(
            decision_routes.decision_service, "dashboard", side_effect=RuntimeError("broken")
        ):
            response = self.client.get("/api/decision-center")
        self.assertEqual(response.status_code, 500)

        with mock.patch.object(
            decision_routes.decision_service, "sync_many", return_value={"rimili": {"ok": True}}
        ):
            response = self.client.post("/api/decision-center/sync", json={"store": "rimili"})
        self.assertEqual(response.status_code, 200)
        self.assertIn("rimili", response.json()["stores"])

        with mock.patch.object(
            self.app.state.container.decision_commands,
            "set_status",
            return_value=DecisionAction(
                fingerprint="rimili:test",
                status=DecisionStatus.COMPLETED,
                updated_at=datetime(2026, 8, 12, tzinfo=UTC),
            ),
        ):
            response = self.client.post(
                "/api/decision-center/status",
                json={"fingerprint": "rimili:test", "status": "completed"},
            )
        self.assertEqual(response.status_code, 200)
        with mock.patch.object(
            self.app.state.container.decision_commands,
            "set_status",
            side_effect=ValueError("bad status"),
        ):
            response = self.client.post(
                "/api/decision-center/status",
                json={"fingerprint": "rimili:test", "status": "completed"},
            )
        self.assertEqual(response.status_code, 400)

        rnp_payload = {
            "ok": True,
            "marketplace": "WB",
            "store": "rimili",
            "items": [],
            "summary": {},
        }
        with mock.patch.object(rnp_routes.rnp_service, "dashboard", return_value=rnp_payload):
            response = self.client.get("/api/rnp?store=rimili&month=2026-08")
        self.assertEqual(response.status_code, 200)
        with mock.patch.object(rnp_routes.rnp_service, "dashboard", side_effect=ValueError("bad")):
            response = self.client.get("/api/rnp?store=rimili&month=bad")
        self.assertEqual(response.status_code, 400)
        with mock.patch.object(rnp_routes.rnp_service, "dashboard", side_effect=RuntimeError("db")):
            response = self.client.get("/api/rnp?store=rimili&month=2026-08")
        self.assertEqual(response.status_code, 500)

        with (
            mock.patch.object(rnp_routes.rnp_service, "sales_lookback_days", return_value=8),
            mock.patch.object(rnp_routes.sales_service, "sync_store", return_value={"ok": True}),
            mock.patch.object(rnp_routes.rnp_service, "sync_metrics", return_value={"ok": True}),
        ):
            response = self.client.post(
                "/api/rnp/sync",
                json={"store": "rimili", "month": "2026-08", "marketplace": "WB"},
            )
        self.assertEqual(response.status_code, 200)

        with mock.patch.object(
            self.app.state.container.rnp_commands,
            "save_strategy",
            return_value=RnpStrategy(
                store_slug="rimili",
                marketplace=Marketplace.WB,
                article="949558341",
                strategy="growth",
                date_from="2026-08-01",
                date_to="2026-08-31",
                updated_by="Unit Admin",
                updated_at=datetime(2026, 8, 12, tzinfo=UTC),
            ),
        ):
            response = self.client.post(
                "/api/rnp/strategy",
                json={
                    "store": "rimili",
                    "marketplace": "WB",
                    "article": "949558341",
                    "strategy": "growth",
                    "date_from": "2026-08-01",
                    "date_to": "2026-08-31",
                },
            )
        self.assertEqual(response.status_code, 200)
        with mock.patch.object(
            self.app.state.container.rnp_commands,
            "add_action",
            return_value=RnpAction(
                id=1,
                article="949558341",
                action_date="2026-08-12",
                note="Done",
                user_name="Unit Admin",
                created_at=datetime(2026, 8, 12, tzinfo=UTC),
            ),
        ):
            response = self.client.post(
                "/api/rnp/action",
                json={
                    "store": "rimili",
                    "marketplace": "WB",
                    "article": "949558341",
                    "action_date": "2026-08-12",
                    "note": "Done",
                },
            )
        self.assertEqual(response.status_code, 200)

    def test_sales_api_contracts(self) -> None:
        with mock.patch.object(
            sales_overview.sales_service,
            "dashboard",
            return_value={"ok": True, "daily": [], "summary": {}},
        ):
            response = self.client.get(
                "/api/sales?store=rimili&marketplace=WB&date_from=2026-08-01&date_to=2026-08-12"
            )
        self.assertEqual(response.status_code, 200)
        with mock.patch.object(
            wb_funnel_orders,
            "dashboard",
            return_value={"ok": True, "marketplace": "WB", "series": []},
        ):
            response = self.client.get(
                "/api/sales/wb-funnel-orders?store=rimili&date_from=2026-08-01&date_to=2026-08-12"
            )
        self.assertEqual(response.status_code, 200)
        with mock.patch.object(
            wb_funnel_orders,
            "dashboard",
            return_value={"ok": True, "marketplace": "WB", "series": []},
        ):
            self.assertEqual(self.client.get("/api/sales/wb-funnel-orders").status_code, 200)
        with mock.patch.object(wb_funnel_orders, "dashboard", side_effect=ValueError("bad period")):
            self.assertEqual(self.client.get("/api/sales/wb-funnel-orders?store=rimili").status_code, 400)
        with mock.patch.object(
            sales_overview.sales_service, "dashboard", side_effect=ValueError("bad period")
        ):
            response = self.client.get("/api/sales?store=rimili")
        self.assertEqual(response.status_code, 400)
        with mock.patch.object(sales_overview.sales_service, "dashboard", side_effect=RuntimeError("db")):
            response = self.client.get("/api/sales?store=rimili")
        self.assertEqual(response.status_code, 500)

        workbook = b"PK\x03\x04unit"
        with mock.patch.object(sales_overview.sales_service, "export_xlsx", return_value=workbook):
            response = self.client.get(
                "/sales/orders.xlsx?store=rimili&marketplace=WB&date_from=2026-08-01&date_to=2026-08-12"
            )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content, workbook)
        with mock.patch.object(
            sales_overview.sales_service, "export_xlsx", side_effect=ValueError("bad period")
        ):
            response = self.client.get("/sales/orders.xlsx?store=rimili")
        self.assertEqual(response.status_code, 400)

    def test_authorization_and_system_failures(self) -> None:
        self.authentication_patch.stop()
        anonymous = mock.patch.object(self.app.state.container.identity, "user_for_token", return_value=None)
        anonymous.start()
        try:
            html_response = self.client.get("/stock", follow_redirects=False)
            json_response = self.client.get("/api/sales", headers={"accept": "application/json"})
            mutation_response = self.client.post("/api/rnp/sync", json={})
        finally:
            anonymous.stop()
            self.authentication_patch.start()
        self.assertEqual(html_response.status_code, 303)
        self.assertEqual(json_response.status_code, 401)
        self.assertEqual(mutation_response.status_code, 401)

        with mock.patch.object(
            self.app.state.container.health,
            "readiness",
            return_value=ReadinessStatus(status="unavailable", database="unavailable"),
        ):
            response = self.client.get("/readyz")
        self.assertEqual(response.status_code, 503)
        self.assertEqual(self.client.get("/healthz").status_code, 200)

    def test_invalid_json_and_forbidden_store_paths(self) -> None:
        limited_user = dict(self.user, role="user", store_slugs=["tris"])
        with mock.patch.object(
            self.app.state.container.identity, "user_for_token", return_value=limited_user
        ):
            stock_page = self.client.get("/stock")
            self.assertEqual(stock_page.status_code, 200)
            self.assertNotIn('id="sync-products-btn"', stock_page.text)
            self.assertEqual(self.client.post("/admin/sync-stock").status_code, 403)
            self.assertEqual(self.client.get("/api/decision-center?store=rimili").status_code, 403)
            self.assertEqual(
                self.client.post("/api/decision-center/sync", json={"store": "rimili"}).status_code,
                403,
            )
            self.assertEqual(
                self.client.post(
                    "/api/decision-center/status",
                    json={"fingerprint": "rimili:test", "status": "completed"},
                ).status_code,
                403,
            )
            self.assertEqual(self.client.get("/api/sales?store=rimili").status_code, 403)
            self.assertEqual(self.client.get("/api/sales/wb-funnel-orders?store=rimili").status_code, 403)
            self.assertEqual(self.client.get("/api/rnp?store=rimili").status_code, 403)
            self.assertEqual(
                self.client.post(
                    "/api/rnp/sync",
                    json={"store": "rimili", "marketplace": "WB", "month": "2026-08"},
                ).status_code,
                403,
            )
        bad_body = "[not-json"
        headers = {"content-type": "application/json"}
        self.assertEqual(
            self.client.post("/api/rnp/strategy", content=bad_body, headers=headers).status_code,
            422,
        )
        self.assertEqual(
            self.client.post("/api/rnp/action", content=json.dumps([]), headers=headers).status_code,
            422,
        )


if __name__ == "__main__":
    unittest.main()
