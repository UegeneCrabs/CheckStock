import io
import unittest
from datetime import date
from unittest import mock

import openpyxl

from app import stock_cost_report, stock_cost_report_export
from app.web.routers import stock_cost_report as stock_cost_report_routes


class StockCostReportTests(unittest.TestCase):
    def test_summary_separates_fbs_transfers_and_calculates_purchase_cost(self) -> None:
        common = {
            "store_slug": "rimili",
            "source_type": "manual",
            "source_name": None,
            "sheet_url": None,
            "from_fulfillment": None,
            "to_fulfillment": None,
            "note": None,
            "user_name": "Test User",
            "created_at": "2026-08-18T10:00:00+00:00",
        }
        operations = [
            common
            | {
                "id": 1,
                "kind": "delivery",
                "from_marketplace": None,
                "to_marketplace": "WB",
                "is_fbs_transfer": 0,
                "items": [{"article": "A", "barcode": "B", "name": "Товар", "quantity": 2}],
            },
            common
            | {
                "id": 2,
                "kind": "transfer",
                "from_marketplace": "WB",
                "to_marketplace": "OZON",
                "is_fbs_transfer": 0,
                "items": [{"article": "A", "barcode": "B", "name": "Товар", "quantity": 1}],
            },
            common
            | {
                "id": 3,
                "kind": "shipment",
                "from_marketplace": "WB",
                "to_marketplace": None,
                "is_fbs_transfer": 0,
                "items": [{"article": "A", "barcode": "B", "name": "Товар", "quantity": 4}],
            },
            common
            | {
                "id": 4,
                "kind": "shipment",
                "from_marketplace": "WB",
                "to_marketplace": None,
                "is_fbs_transfer": 1,
                "items": [{"article": "A", "barcode": "B", "name": "Товар", "quantity": 3}],
            },
            common
            | {
                "id": 5,
                "kind": "fbs_transfer",
                "from_marketplace": "WB",
                "to_marketplace": "WB",
                "is_fbs_transfer": 0,
                "items": [{"article": "A", "barcode": "B", "name": "Товар", "quantity": 2}],
            },
            common
            | {
                "id": 6,
                "kind": "transfer_dispatch",
                "from_marketplace": "WB",
                "to_marketplace": "OZON",
                "is_fbs_transfer": 0,
                "items": [{"article": "A", "barcode": "B", "name": "Товар", "quantity": 2}],
            },
            common
            | {
                "id": 7,
                "kind": "transfer_receive",
                "from_marketplace": "WB",
                "to_marketplace": "OZON",
                "is_fbs_transfer": 0,
                "items": [{"article": "A", "barcode": "B", "name": "Товар", "quantity": 2}],
            },
            common
            | {
                "id": 8,
                "kind": "transfer_cancel",
                "from_marketplace": "OZON",
                "to_marketplace": "WB",
                "is_fbs_transfer": 0,
                "items": [{"article": "A", "barcode": "B", "name": "Товар", "quantity": 1}],
            },
        ]
        snapshots = [
            {
                "store_slug": "rimili",
                "marketplace": "WB",
                "article": "A",
                "barcode": "B",
                "name": "Товар",
                "day": "2026-08-16",
                "quantity": 10,
            },
            {
                "store_slug": "rimili",
                "marketplace": "WB",
                "article": "A",
                "barcode": "B",
                "name": "Товар",
                "day": "2026-08-23",
                "quantity": 8,
            },
        ]
        coverage = {
            ("rimili", "WB", "2026-08-16"),
            ("rimili", "WB", "2026-08-23"),
        }
        with (
            mock.patch.object(
                stock_cost_report.db,
                "get_purchase_price_rows",
                return_value=[
                    {
                        "store_slug": "rimili",
                        "article": "A",
                        "barcode": "B",
                        "name": "Товар",
                        "purchase_price": 10,
                    }
                ],
            ),
            mock.patch.object(
                stock_cost_report.db,
                "get_operations_with_items_for_period",
                return_value=operations,
            ),
            mock.patch.object(
                stock_cost_report.db,
                "get_fbs_sales_for_period",
                return_value=[
                    {
                        "store_slug": "rimili",
                        "marketplace": "WB",
                        "article": "A",
                        "barcode": "B",
                        "name": "Товар",
                        "quantity": 5,
                    }
                ],
            ),
            mock.patch.object(
                stock_cost_report.db,
                "get_fbs_stock_snapshots",
                return_value=(snapshots, coverage),
            ),
        ):
            report = stock_cost_report.build_report(
                ("rimili",),
                date(2026, 8, 17),
                date(2026, 8, 23),
                ("WB", "OZON"),
            )

        summary = {(row["store_slug"], row["marketplace"]): row for row in report["summary"]}
        wb = summary[("rimili", "WB")]
        ozon = summary[("rimili", "OZON")]
        self.assertEqual((wb["deliveries"]["units"], wb["deliveries"]["cost"]), (2, 20.0))
        self.assertEqual((wb["moved_out"]["units"], wb["moved_in"]["units"]), (3, 6))
        self.assertEqual((wb["shipped"]["units"], wb["shipped"]["cost"]), (4, 40.0))
        self.assertEqual((wb["fbs_sales"]["units"], wb["fbs_sales"]["cost"]), (7, 70.0))
        self.assertEqual((wb["fbs_actual_sales"]["units"], wb["fbs_actual_sales"]["cost"]), (5, 50.0))
        self.assertEqual(ozon["moved_in"]["units"], 3)
        self.assertEqual(len(stock_cost_report.operations_for_view(report, "transfers")), 4)
        self.assertTrue(wb["fbs_formula"]["available"])
        self.assertEqual(wb["fbs_formula"]["metric"]["units"], 7)

        sales_table = stock_cost_report_routes._fbs_sales_table(report)
        self.assertIn('class="cost-sales-table data-table"', sales_table)
        self.assertIn("<th>Магазин</th><th>Маркетплейс</th>", sales_table)

        content, filename = stock_cost_report_export.build_xlsx(report, "summary")
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        self.assertEqual(workbook.sheetnames, ["Операции", "По артикулам"])
        self.assertEqual(workbook["Операции"]["C1"].value, "Маркетплейс")
        self.assertEqual(workbook["По артикулам"]["C1"].value, "Маркетплейс")
        self.assertIn("dvizhenie_zc_2026-08-17_2026-08-23", filename)
        self.assertGreater(workbook["Операции"].max_row, 2)
        self.assertGreater(workbook["По артикулам"].max_row, 2)


if __name__ == "__main__":
    unittest.main()
